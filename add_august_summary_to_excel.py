import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def add_august_summary_to_excel():
    """Add August metrics summary to the Excel file with proper formatting"""
    
    try:
        file_path = "August_Export_SD_2_Sept_updated.xlsx"
        print(f"Adding August metrics summary to: {file_path}")
        
        # Load the workbook
        workbook = load_workbook(file_path)
        
        # Create or get the August Summary sheet
        if 'August_Summary' in workbook.sheetnames:
            workbook.remove(workbook['August_Summary'])
        
        summary_sheet = workbook.create_sheet('August_Summary')
        
        # Define styles
        title_font = Font(bold=True, size=16, color="FFFFFF")
        header_font = Font(bold=True, size=12, color="FFFFFF")
        data_font = Font(size=11)
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        data_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Add title
        summary_sheet.merge_cells('A1:D1')
        title_cell = summary_sheet['A1']
        title_cell.value = "AUGUST METRICS SUMMARY (Total across August)"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add headers
        headers = [
            'Metric',
            'Value',
            'Formula',
            'Description'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = summary_sheet.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data rows
        data_rows = [
            {
                'Metric': 'Total Students in August',
                'Value': 374,
                'Formula': '=COUNTA(August!A:A)-1',
                'Description': 'Total number of students in August dataset'
            },
            {
                'Metric': 'Average VWE modules commenced per student',
                'Value': 1.60,
                'Formula': '=AVERAGE(August!K:K)',
                'Description': 'Average Virtual Work Experience modules started per student'
            },
            {
                'Metric': 'Average industry-based modules completed per student',
                'Value': 4.13,
                'Formula': '=AVERAGE(Industry_Module_Count)',
                'Description': 'Average number of industry modules completed per student'
            },
            {
                'Metric': 'Average modules engaged with per session per student',
                'Value': 2.54,
                'Formula': '=AVERAGE(Modules_Per_Session)',
                'Description': 'Average modules engaged per web session per student'
            }
        ]
        
        for row_idx, data in enumerate(data_rows, 4):
            for col_idx, (key, value) in enumerate(data.items(), 1):
                cell = summary_sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                if col_idx == 1:  # Metric column
                    cell.fill = data_fill
                    cell.font = Font(bold=True, size=11)
                elif col_idx == 2:  # Value column
                    cell.font = Font(bold=True, size=11, color="0000FF")
                elif col_idx == 3:  # Formula column
                    cell.font = Font(size=10, color="008000")
                elif col_idx == 4:  # Description column
                    cell.font = Font(size=10, color="666666")
        
        # Add detailed breakdown section
        print("Adding detailed breakdown section...")
        
        # VWE Breakdown
        row_start = 10
        summary_sheet.cell(row=row_start, column=1, value="VWE MODULES BREAKDOWN").font = Font(bold=True, size=14)
        summary_sheet.merge_cells(f'A{row_start}:D{row_start}')
        
        vwe_breakdown = [
            ['VWE Level', 'Count', 'Percentage', 'Formula'],
            ['1 Module', 103, '53.6%', '=COUNTIF(August!K:K,1)'],
            ['2 Modules', 66, '34.4%', '=COUNTIF(August!K:K,2)'],
            ['3 Modules', 19, '9.9%', '=COUNTIF(August!K:K,3)'],
            ['4 Modules', 4, '2.1%', '=COUNTIF(August!K:K,4)'],
            ['Total with VWE', 192, '100.0%', '=COUNTA(August!K:K)']
        ]
        
        for row_idx, row_data in enumerate(vwe_breakdown, row_start + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == row_start + 1:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.font = data_font
                    if col_idx == 1:  # VWE Level column
                        cell.fill = data_fill
                        cell.font = Font(bold=True)
        
        # Industry Modules Breakdown
        row_start = 18
        summary_sheet.cell(row=row_start, column=1, value="INDUSTRY MODULES BREAKDOWN").font = Font(bold=True, size=14)
        summary_sheet.merge_cells(f'A{row_start}:D{row_start}')
        
        industry_breakdown = [
            ['Module Count Range', 'Students', 'Percentage', 'Formula'],
            ['1-3 Modules', 168, '50.5%', '=COUNTIFS(Industry_Module_Count,"<=3")'],
            ['4-6 Modules', 117, '35.1%', '=COUNTIFS(Industry_Module_Count,">=4",Industry_Module_Count,"<=6")'],
            ['7+ Modules', 48, '14.4%', '=COUNTIFS(Industry_Module_Count,">=7")'],
            ['Total with Industry Data', 333, '100.0%', '=COUNTA(August!J:J)']
        ]
        
        for row_idx, row_data in enumerate(industry_breakdown, row_start + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == row_start + 1:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.font = data_font
                    if col_idx == 1:  # Module Count Range column
                        cell.fill = data_fill
                        cell.font = Font(bold=True)
        
        # Engagement Breakdown
        row_start = 26
        summary_sheet.cell(row=row_start, column=1, value="ENGAGEMENT PER SESSION BREAKDOWN").font = Font(bold=True, size=14)
        summary_sheet.merge_cells(f'A{row_start}:D{row_start}')
        
        engagement_breakdown = [
            ['Modules per Session', 'Students', 'Percentage', 'Formula'],
            ['0-2 Modules', 135, '49.6%', '=COUNTIFS(Modules_Per_Session,"<=2")'],
            ['3-5 Modules', 98, '36.0%', '=COUNTIFS(Modules_Per_Session,">=3",Modules_Per_Session,"<=5")'],
            ['6+ Modules', 39, '14.4%', '=COUNTIFS(Modules_Per_Session,">=6")'],
            ['Total with Engagement Data', 272, '100.0%', '=COUNTA(August!E:E)']
        ]
        
        for row_idx, row_data in enumerate(engagement_breakdown, row_start + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == row_start + 1:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.font = data_font
                    if col_idx == 1:  # Modules per Session column
                        cell.fill = data_fill
                        cell.font = Font(bold=True)
        
        # Add notes section
        row_start = 35
        summary_sheet.cell(row=row_start, column=1, value="NOTES & DEFINITIONS").font = Font(bold=True, size=14, color="FFFFFF")
        summary_sheet.merge_cells(f'A{row_start}:D{row_start}')
        summary_sheet.cell(row=row_start, column=1).fill = title_fill
        
        notes = [
            ['Term', 'Definition', 'Data Source', 'Calculation Method'],
            ['VWE', 'Virtual Work Experience modules', 'Column K in August sheet', 'Direct average of numeric values'],
            ['Industry Modules', 'Industry preference selections', 'Column J in August sheet', 'Count of pipe-separated values'],
            ['Modules per Session', 'Engagement types per web session', 'Person tag + Web sessions', 'Engagement count ÷ Session count'],
            ['Web Sessions', 'Number of web sessions per student', 'Column C in August sheet', 'Direct count from data']
        ]
        
        for row_idx, row_data in enumerate(notes, row_start + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == row_start + 1:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.font = data_font
                    if col_idx == 1:  # Term column
                        cell.fill = data_fill
                        cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in summary_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            summary_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        workbook.save(file_path)
        print(f"August summary sheet added successfully to: {file_path}")
        
        # Display what was added
        print(f"\nAdded to August_Summary sheet:")
        print(f"- Main metrics with formulas")
        print(f"- VWE modules breakdown")
        print(f"- Industry modules breakdown") 
        print(f"- Engagement per session breakdown")
        print(f"- Notes and definitions")
        print(f"- Professional formatting and styling")
        
        return True
        
    except Exception as e:
        print(f"Error adding August summary: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = add_august_summary_to_excel()
    if success:
        print(f"\nAugust summary sheet creation completed successfully!")
    else:
        print(f"\nFailed to create August summary sheet.")
