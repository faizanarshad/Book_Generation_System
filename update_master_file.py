import pandas as pd
import numpy as np
from openpyxl import load_workbook

def update_master_file():
    """Update the master file directly by adding '1' in column T when 'Career Profiling Engaged' appears in column E"""
    
    try:
        # Read the Excel file
        file_path = "August Export_SD 2 Sept_modified.xlsx"
        print(f"Reading data from: {file_path}")
        
        # Load the workbook
        workbook = load_workbook(file_path)
        
        # Check if August sheet exists
        if 'August' not in workbook.sheetnames:
            print("Error: 'August' sheet not found!")
            return
        
        # Get the August worksheet
        august_worksheet = workbook['August']
        print(f"August sheet loaded successfully")
        
        # Find the column T (20th column, index 19)
        # First, let's check if we need to add a new column
        max_col = august_worksheet.max_column
        
        if max_col < 20:
            # Add new column T
            print("Adding new column T (Career Profiling Flag)")
            # Insert column T at position 20
            august_worksheet.insert_cols(20)
            # Add header
            august_worksheet.cell(row=1, column=20, value="Career_Profiling_Flag")
        else:
            # Use existing column T
            print("Using existing column T")
            # Update header
            august_worksheet.cell(row=1, column=20, value="Career_Profiling_Flag")
        
        # Column E is the 5th column (index 4)
        person_tag_col = 5
        
        # Initialize counter
        updated_count = 0
        
        # Process each row starting from row 2 (skip header)
        for row in range(2, august_worksheet.max_row + 1):
            person_tag_value = august_worksheet.cell(row=row, column=person_tag_col).value
            
            if person_tag_value and "Career Profiling Engaged" in str(person_tag_value):
                # Set column T to 1
                august_worksheet.cell(row=row, column=20, value=1)
                updated_count += 1
            else:
                # Set column T to 0
                august_worksheet.cell(row=row, column=20, value=0)
        
        print(f"Updated {updated_count} rows with '1' in column T")
        
        # Save the updated master file
        workbook.save(file_path)
        print(f"Master file updated and saved: {file_path}")
        
        # Verify the update by reading the file again
        verification_df = pd.read_excel(file_path, sheet_name='August')
        verification_count = verification_df.iloc[:, 19].sum()  # Column T (index 19)
        print(f"Verification: {verification_count} rows have '1' in column T")
        
        # Show sample of updated data
        print("\nSample of updated data:")
        updated_rows = verification_df[verification_df.iloc[:, 19] == 1]
        if not updated_rows.empty:
            print(updated_rows[['First name', 'Person tag', verification_df.columns[19]]].head())
        
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
    except Exception as e:
        print(f"Error processing file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    update_master_file()
