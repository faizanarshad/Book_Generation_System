#!/usr/bin/env python3
"""
August Data Analysis
This script analyzes August data to answer specific questions about users,
login counts, average time, and creates pivot tables by Year group and International status.
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import numpy as np

class AugustAnalysis:
    def __init__(self, file_path):
        self.file_path = file_path
        self.august_data = None
        self.analysis_results = {}
        
    def load_august_data(self):
        """Load August data from the Excel file"""
        try:
            self.august_data = pd.read_excel(self.file_path, sheet_name='August')
            print(f"August data loaded: {len(self.august_data)} rows")
            print(f"August columns: {list(self.august_data.columns)}")
            return True
        except Exception as e:
            print(f"Error loading August data: {e}")
            return False
    
    def analyze_basic_stats(self):
        """Analyze basic statistics from August data"""
        if self.august_data is None:
            print("No August data loaded.")
            return False
            
        try:
            # Total users (unique emails)
            total_users = self.august_data['Email'].nunique()
            
            # Total login count (sum of all login counts)
            # Note: August data has 'Web sessions' instead of 'Login Count'
            login_count_column = 'Web sessions' if 'Web sessions' in self.august_data.columns else 'Login Count'
            total_login_count = self.august_data[login_count_column].sum()
            
            # Average time spent per session
            avg_time_column = 'Avg Login Time'
            avg_time_per_session = self.august_data[avg_time_column].mean()
            
            self.analysis_results['basic_stats'] = {
                'total_users': total_users,
                'total_login_count': total_login_count,
                'avg_time_per_session': avg_time_per_session,
                'login_count_column': login_count_column
            }
            
            print(f"Total Users: {total_users}")
            print(f"Total Login Count (Web Sessions): {total_login_count}")
            print(f"Average Time per Session: {avg_time_per_session:.2f} seconds")
            
            return True
            
        except Exception as e:
            print(f"Error analyzing basic stats: {e}")
            return False
    
    def create_pivot_table(self):
        """Create pivot table for Login Count by Year group and International status"""
        if self.august_data is None:
            print("No August data loaded.")
            return None
            
        try:
            # Prepare data for pivot table
            pivot_data = self.august_data.copy()
            
            # Clean and standardize column names
            year_column = 'Course Year'
            international_column = 'International Status'
            login_column = 'Web sessions' if 'Web sessions' in self.august_data.columns else 'Login Count'
            
            # Clean the data
            pivot_data[year_column] = pivot_data[year_column].fillna('(blank)')
            pivot_data[international_column] = pivot_data[international_column].fillna('(blank)')
            pivot_data[login_column] = pivot_data[login_column].fillna(0)
            
            # Standardize year values (remove pipe characters and clean)
            pivot_data[year_column] = pivot_data[year_column].astype(str).str.strip()
            pivot_data[year_column] = pivot_data[year_column].str.replace("'|", "").str.replace("|'", "").str.replace("|", "")
            pivot_data[year_column] = pivot_data[year_column].replace({
                '1st Year': '1st Year',
                '2nd Year': '2nd Year', 
                '3rd Year': '3rd Year',
                '4th Year': '4th Year',
                '5th Year': '5th Year',
                'nan': '(blank)',
                '': '(blank)'
            })
            
            # Standardize international status (remove pipe characters and clean)
            pivot_data[international_column] = pivot_data[international_column].astype(str).str.strip()
            pivot_data[international_column] = pivot_data[international_column].str.replace("'|", "").str.replace("|'", "").str.replace("|", "")
            pivot_data[international_column] = pivot_data[international_column].replace({
                'Domestic': 'Domestic',
                'International': 'International',
                'nan': '(blank)',
                '': '(blank)'
            })
            
            # Create pivot table
            pivot_table = pd.pivot_table(
                pivot_data,
                values=login_column,
                index=international_column,
                columns=year_column,
                aggfunc='sum',
                fill_value=0,
                margins=True,
                margins_name='Grand Total'
            )
            
            # Reorder columns to match the requested format
            year_order = ['1st Year', '2nd Year', '3rd Year', '4th Year', '5th Year', '(blank)', 'Grand Total']
            existing_columns = [col for col in year_order if col in pivot_table.columns]
            pivot_table = pivot_table[existing_columns]
            
            # Reorder rows to match the requested format
            row_order = ['Domestic', 'International', '(blank)', 'Grand Total']
            existing_rows = [row for row in row_order if row in pivot_table.index]
            pivot_table = pivot_table.reindex(existing_rows)
            
            self.analysis_results['pivot_table'] = pivot_table
            
            print("Pivot table created successfully:")
            print(pivot_table)
            
            return pivot_table
            
        except Exception as e:
            print(f"Error creating pivot table: {e}")
            return None
    
    def create_excel_report(self, output_filename="August_Analysis_Report.xlsx"):
        """Create Excel report with analysis results"""
        try:
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create Summary sheet
            self.create_summary_sheet(wb)
            
            # Create Pivot Table sheet
            self.create_pivot_sheet(wb)
            
            # Create Raw Data sheet
            self.create_raw_data_sheet(wb)
            
            # Save workbook
            wb.save(output_filename)
            print(f"Excel report saved as: {output_filename}")
            return True
            
        except Exception as e:
            print(f"Error creating Excel report: {e}")
            return False
    
    def create_summary_sheet(self, wb):
        """Create summary statistics sheet"""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws.merge_cells('A1:B1')
        ws['A1'] = "AUGUST DATA ANALYSIS SUMMARY"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Summary data
        if 'basic_stats' in self.analysis_results:
            stats = self.analysis_results['basic_stats']
            
            summary_data = [
                ["Metric", "Value"],
                ["Total Users", stats['total_users']],
                ["Total Login Count (Web Sessions)", stats['total_login_count']],
                ["Average Time per Session (seconds)", round(stats['avg_time_per_session'], 2)],
                ["Average Time per Session (minutes)", round(stats['avg_time_per_session'] / 60, 2)],
                ["", ""],
                ["Data Source", "August sheet from August Export_SD 2 Sept.xlsx"],
                ["Login Count Column", stats['login_count_column']],
                ["Analysis Date", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")]
            ]
            
            for row_idx, row_data in enumerate(summary_data, 3):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 3:  # Header row
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    elif col_idx == 1 and value in ["", "Data Source", "Login Count Column", "Analysis Date"]:
                        if value == "":
                            pass  # Empty row
                        else:
                            cell.font = Font(italic=True)
        
        # Auto-adjust columns
        self.auto_adjust_columns(ws)
    
    def create_pivot_sheet(self, wb):
        """Create pivot table sheet"""
        ws = wb.create_sheet("Pivot Table")
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = "LOGIN COUNT BY YEAR GROUP AND INTERNATIONAL STATUS"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        if 'pivot_table' in self.analysis_results:
            pivot_table = self.analysis_results['pivot_table']
            
            # Add pivot table data
            for row_idx, (index, row) in enumerate(pivot_table.iterrows(), 3):
                # Add row label
                ws.cell(row=row_idx, column=1, value=index)
                
                # Add data values
                for col_idx, value in enumerate(row, 2):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Add column headers
            for col_idx, col_name in enumerate(pivot_table.columns, 2):
                ws.cell(row=2, column=col_idx, value=col_name)
            
            # Add row header
            ws.cell(row=2, column=1, value="International Status")
            
            # Style headers
            for col in range(1, len(pivot_table.columns) + 2):
                cell = ws.cell(row=2, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Style row labels
            for row in range(3, len(pivot_table) + 3):
                cell = ws.cell(row=row, column=1)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # Auto-adjust columns
        self.auto_adjust_columns(ws)
    
    def create_raw_data_sheet(self, wb):
        """Create raw data sheet"""
        ws = wb.create_sheet("August Raw Data")
        
        # Add raw data
        for r in dataframe_to_rows(self.august_data, index=False, header=True):
            ws.append(r)
        
        # Style header
        self.style_header(ws, 1)
        
        # Auto-adjust columns
        self.auto_adjust_columns(ws)
    
    def style_header(self, ws, row):
        """Style the header row"""
        for cell in ws[row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
    
    def auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def print_analysis_results(self):
        """Print analysis results in a formatted way"""
        print("\n" + "="*80)
        print("AUGUST DATA ANALYSIS RESULTS")
        print("="*80)
        
        if 'basic_stats' in self.analysis_results:
            stats = self.analysis_results['basic_stats']
            print(f"\nTOTAL:")
            print(f"  {stats['total_users']} Users")
            print(f"  {stats['total_login_count']} Login Count (Web Sessions)")
            print(f"\nAverage time spent on platform per session:")
            print(f"  {stats['avg_time_per_session']:.2f} seconds")
            print(f"  {stats['avg_time_per_session']/60:.2f} minutes")
        
        if 'pivot_table' in self.analysis_results:
            print(f"\nPIVOT TABLE - Sum of Login Count by Year Group and International Status:")
            print("-" * 80)
            pivot_table = self.analysis_results['pivot_table']
            print(pivot_table.to_string())

def main():
    print("AUGUST DATA ANALYSIS")
    print("="*50)
    
    # Initialize analysis
    analysis = AugustAnalysis("August Export_SD 2 Sept.xlsx")
    
    # Load August data
    if not analysis.load_august_data():
        return
    
    # Analyze basic statistics
    if not analysis.analyze_basic_stats():
        return
    
    # Create pivot table
    pivot_table = analysis.create_pivot_table()
    if pivot_table is None:
        return
    
    # Print results
    analysis.print_analysis_results()
    
    # Create Excel report
    if analysis.create_excel_report():
        print("\n" + "="*60)
        print("ANALYSIS COMPLETE")
        print("="*60)
        print("Excel report saved as: August_Analysis_Report.xlsx")
        print("This file contains:")
        print("1. Summary sheet - Basic statistics and totals")
        print("2. Pivot Table sheet - Login count by Year group and International status")
        print("3. August Raw Data sheet - Complete August dataset")

if __name__ == "__main__":
    main()
