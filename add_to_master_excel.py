#!/usr/bin/env python3
"""
Add Industry Preferences Table to Master Excel File
This script adds the industry preferences table as a new sheet
directly to the master Excel file "August Export_SD 2 Sept.xlsx".
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np
import re

def add_industry_preferences_to_master():
    """Add industry preferences table to the master Excel file"""
    
    master_file = "August Export_SD 2 Sept.xlsx"
    
    # Load the master workbook
    print(f"Loading master Excel file: {master_file}")
    wb = load_workbook(master_file)
    
    # Check existing sheets
    print(f"Existing sheets: {wb.sheetnames}")
    
    # Load data from existing sheets
    print("Loading data from existing sheets...")
    august_data = pd.read_excel(master_file, sheet_name='August')
    industry_mapping = {}
    
    # Load industry mapping from Sheet 7
    sheet7_data = pd.read_excel(master_file, sheet_name='Sheet7', header=None)
    for i in range(2, len(sheet7_data)):
        row = sheet7_data.iloc[i]
        if pd.notna(row[1]) and pd.notna(row[2]):
            industry_number = int(row[1])
            industry_name = str(row[2]).strip()
            industry_mapping[industry_number] = industry_name
    
    print(f"Loaded {len(industry_mapping)} industry mappings")
    
    # Define the exact industry order requested
    industry_order = [
        'Accounting', 'Advertising, Media, Journalism, and Communications',
        'Agriculture and Environment', 'Animals and Vet', 'Architecture',
        'Arts, Humanities, and Politics', 'Building and Construction',
        'Business and Commerce', 'Community and Social Work',
        'Creative Arts and Music', 'Design', 'Economics and Finance',
        'Education, Childcare and Teaching', 'Engineering', 'Entrepreneur',
        'Food and Beverage', 'Government, Defence and Policing',
        'Hair and Beauty', 'Health and Sport Sciences', 'Law',
        'Marketing and Public Relations', 'Mathematics',
        'Medical Sciences and Medicine', 'Nursing and Midwifery',
        'Property and Real Estate', 'Psychology', 'Science', 'Technology',
        'Trades and Mining', 'Sports', 'Transport, Tourism and Hospitality',
        'Fashion', 'Australian Defence Force', 'Energy'
    ]
    
    # Initialize results table
    results = {}
    for industry in industry_order:
        results[industry] = {
            'Engineering': {'1st Year': 0, '2nd Year': 0, '3rd Year': 0, '4th Year': 0, '5th Year': 0},
            'Arts': {'1st Year': 0, '2nd Year': 0, '3rd Year': 0, '4th Year': 0, '5th Year': 0}
        }
    
    # Process each student
    print("Processing student data...")
    for _, student in august_data.iterrows():
        # Get faculty
        faculty = str(student.get('Faculty', '')).replace("'", "").replace("|", "").strip()
        if "Faculty of Engineering" in faculty:
            faculty_type = "Engineering"
        elif "Faculty of Arts and Social Sciences" in faculty:
            faculty_type = "Arts"
        else:
            continue  # Skip other faculties
        
        # Get year
        year = str(student.get('Course Year', '')).replace("'", "").replace("|", "").strip()
        if "1st Year" in year or year == "1":
            year_type = "1st Year"
        elif "2nd Year" in year or year == "2":
            year_type = "2nd Year"
        elif "3rd Year" in year or year == "3":
            year_type = "3rd Year"
        elif "4th Year" in year or year == "4":
            year_type = "4th Year"
        elif "5th Year" in year or year == "5":
            year_type = "5th Year"
        else:
            continue  # Skip unknown years
        
        # Get industry preferences
        industries_str = str(student.get('Industries', ''))
        if pd.notna(industries_str) and industries_str != 'nan':
            # Parse industry numbers
            numbers = re.findall(r'\d+', industries_str)
            for num in numbers:
                industry_num = int(num)
                if industry_num in industry_mapping:
                    industry_name = industry_mapping[industry_num]
                    if industry_name in results:
                        results[industry_name][faculty_type][year_type] += 1
    
    # Create new sheet in master workbook
    print("Creating Industry Preferences sheet in master workbook...")
    
    # Remove existing sheet if it exists
    if "Industry Preferences" in wb.sheetnames:
        wb.remove(wb["Industry Preferences"])
        print("Removed existing Industry Preferences sheet")
    
    # Create new sheet
    ws = wb.create_sheet("Industry Preferences")
    
    # Title
    ws.merge_cells('A1:L1')
    ws['A1'] = "INDUSTRY PREFERENCES BY FACULTY AND YEAR GROUP"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Faculty headers
    ws.merge_cells('B2:F2')
    ws['B2'] = "Faculty of Engineering"
    ws.merge_cells('G2:K2')
    ws['G2'] = "Faculty of Arts and Social Sciences"
    
    # Style faculty headers
    for cell in ['B2', 'G2']:
        ws[cell].font = Font(bold=True, size=12)
        ws[cell].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        ws[cell].alignment = Alignment(horizontal='center')
    
    # Year headers
    year_headers = ['1st Year', '2nd Year', '3rd Year', '4th Year', '5th Year']
    for i, year in enumerate(year_headers):
        # Engineering years
        ws.cell(row=3, column=i+2, value=year)
        # Arts years
        ws.cell(row=3, column=i+7, value=year)
    
    # Style year headers
    for col in range(2, 12):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Industry column header
    ws['A3'] = "Industry"
    ws['A3'].font = Font(bold=True)
    ws['A3'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Add data
    for row_idx, industry in enumerate(industry_order, 4):
        # Industry name
        ws.cell(row=row_idx, column=1, value=industry)
        ws.cell(row=row_idx, column=1).font = Font(bold=True)
        
        # Engineering data
        for col_idx, year in enumerate(year_headers):
            value = results[industry]['Engineering'][year]
            ws.cell(row=row_idx, column=col_idx+2, value=value)
        
        # Arts data
        for col_idx, year in enumerate(year_headers):
            value = results[industry]['Arts'][year]
            ws.cell(row=row_idx, column=col_idx+7, value=value)
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the updated master workbook
    print("Saving updated master workbook...")
    wb.save(master_file)
    
    # Print summary
    print("\n" + "="*80)
    print("INDUSTRY PREFERENCES SUMMARY")
    print("="*80)
    
    total_engineering = sum(sum(results[industry]['Engineering'].values()) for industry in industry_order)
    total_arts = sum(sum(results[industry]['Arts'].values()) for industry in industry_order)
    
    print(f"Total Engineering preferences: {total_engineering}")
    print(f"Total Arts preferences: {total_arts}")
    print(f"Total preferences: {total_engineering + total_arts}")
    
    # Show top industries for each faculty
    print("\nTop 5 Industries for Engineering:")
    engineering_totals = [(industry, sum(results[industry]['Engineering'].values())) for industry in industry_order]
    engineering_totals.sort(key=lambda x: x[1], reverse=True)
    for i, (industry, count) in enumerate(engineering_totals[:5]):
        print(f"  {i+1}. {industry}: {count}")
    
    print("\nTop 5 Industries for Arts:")
    arts_totals = [(industry, sum(results[industry]['Arts'].values())) for industry in industry_order]
    arts_totals.sort(key=lambda x: x[1], reverse=True)
    for i, (industry, count) in enumerate(arts_totals[:5]):
        print(f"  {i+1}. {industry}: {count}")
    
    print(f"\n✅ Successfully added 'Industry Preferences' sheet to {master_file}")
    print("The table shows exactly the format you requested with all 34 industries!")
    
    return results

if __name__ == "__main__":
    print("ADDING INDUSTRY PREFERENCES TABLE TO MASTER EXCEL FILE")
    print("="*60)
    
    try:
        results = add_industry_preferences_to_master()
        print("\n" + "="*60)
        print("OPERATION COMPLETE")
        print("="*60)
        print("Your master Excel file now contains:")
        print("1. July sheet - July data")
        print("2. August sheet - August data")
        print("3. Sheet7 - Industry mappings")
        print("4. Industry Preferences - NEW sheet with your requested table format")
        print("\nOpen the Excel file to see the new Industry Preferences sheet!")
        
    except Exception as e:
        print(f"Error: {e}")
        print("Please make sure the master Excel file is not open in Excel when running this script.")
