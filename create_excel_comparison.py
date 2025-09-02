#!/usr/bin/env python3
"""
Create Excel Comparison Workbook with Formulas
This script creates a comprehensive Excel workbook with formulas to automatically
calculate July-August user activity comparisons.
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import numpy as np

class ExcelComparisonCreator:
    def __init__(self, source_file):
        self.source_file = source_file
        self.workbook = None
        self.july_data = None
        self.august_data = None
        
    def load_source_data(self):
        """Load data from the source Excel file"""
        try:
            self.july_data = pd.read_excel(self.source_file, sheet_name='July ')
            self.august_data = pd.read_excel(self.source_file, sheet_name='August')
            print(f"Loaded July data: {len(self.july_data)} rows")
            print(f"Loaded August data: {len(self.august_data)} rows")
            return True
        except Exception as e:
            print(f"Error loading source data: {e}")
            return False
    
    def create_workbook(self):
        """Create a new workbook with multiple sheets"""
        self.workbook = Workbook()
        
        # Remove default sheet
        self.workbook.remove(self.workbook.active)
        
        # Create sheets
        self.create_july_sheet()
        self.create_august_sheet()
        self.create_comparison_sheet()
        self.create_summary_sheet()
        self.create_formulas_guide_sheet()
        
        return True
    
    def create_july_sheet(self):
        """Create July data sheet"""
        ws = self.workbook.create_sheet("July Data")
        
        # Add data
        for r in dataframe_to_rows(self.july_data, index=False, header=True):
            ws.append(r)
        
        # Style the header
        self.style_header(ws, 1)
        
        # Auto-adjust columns
        self.auto_adjust_columns(ws)
        
        print("Created July Data sheet")
    
    def create_august_sheet(self):
        """Create August data sheet"""
        ws = self.workbook.create_sheet("August Data")
        
        # Add data
        for r in dataframe_to_rows(self.august_data, index=False, header=True):
            ws.append(r)
        
        # Style the header
        self.style_header(ws, 1)
        
        # Auto-adjust columns
        self.auto_adjust_columns(ws)
        
        print("Created August Data sheet")
    
    def create_comparison_sheet(self):
        """Create comparison sheet with formulas"""
        ws = self.workbook.create_sheet("July-August Comparison")
        
        # Headers
        headers = [
            "Email", "First Name", 
            "July Login Count", "August Login Count", "Login Count Increase", "Login Count % Change",
            "July Avg Login Time", "August Avg Login Time", "Avg Time Increase (sec)", "Avg Time % Change",
            "July VWE", "August VWE", "VWE Increase", "VWE % Change"
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Style header
        self.style_header(ws, 1)
        
        # Get unique emails from both months
        july_emails = set(self.july_data['Email'].dropna().astype(str))
        august_emails = set(self.august_data['Email'].dropna().astype(str))
        existing_emails = july_emails.intersection(august_emails)
        
        # Add comparison data with formulas
        row = 2
        for email in sorted(existing_emails):
            # Email and First Name
            ws.cell(row=row, column=1, value=email)
            
            # Get August first name
            august_user = self.august_data[self.august_data['Email'].astype(str) == email]
            if len(august_user) > 0:
                ws.cell(row=row, column=2, value=august_user.iloc[0].get('First name', ''))
            
            # July Login Count (using VLOOKUP)
            ws.cell(row=row, column=3, value=f'=VLOOKUP(A{row},\'July Data\'!B:C,2,FALSE)')
            
            # August Login Count (using VLOOKUP)
            ws.cell(row=row, column=4, value=f'=VLOOKUP(A{row},\'August Data\'!B:C,2,FALSE)')
            
            # Login Count Increase
            ws.cell(row=row, column=5, value=f'=D{row}-C{row}')
            
            # Login Count % Change
            ws.cell(row=row, column=6, value=f'=IF(C{row}=0,0,(D{row}-C{row})/C{row}*100)')
            
            # July Avg Login Time
            ws.cell(row=row, column=7, value=f'=VLOOKUP(A{row},\'July Data\'!B:D,3,FALSE)')
            
            # August Avg Login Time
            ws.cell(row=row, column=8, value=f'=VLOOKUP(A{row},\'August Data\'!B:D,3,FALSE)')
            
            # Avg Time Increase
            ws.cell(row=row, column=9, value=f'=H{row}-G{row}')
            
            # Avg Time % Change
            ws.cell(row=row, column=10, value=f'=IF(G{row}=0,0,(H{row}-G{row})/G{row}*100)')
            
            # July VWE
            ws.cell(row=row, column=11, value=f'=VLOOKUP(A{row},\'July Data\'!B:K,10,FALSE)')
            
            # August VWE
            ws.cell(row=row, column=12, value=f'=VLOOKUP(A{row},\'August Data\'!B:K,10,FALSE)')
            
            # VWE Increase
            ws.cell(row=row, column=13, value=f'=L{row}-K{row}')
            
            # VWE % Change
            ws.cell(row=row, column=14, value=f'=IF(K{row}=0,0,(L{row}-K{row})/K{row}*100)')
            
            row += 1
        
        # Auto-adjust columns
        self.auto_adjust_columns(ws)
        
        print(f"Created July-August Comparison sheet with {len(existing_emails)} users")
    
    def create_summary_sheet(self):
        """Create summary statistics sheet with formulas"""
        ws = self.workbook.create_sheet("Summary Statistics")
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = "JULY-AUGUST USER ACTIVITY SUMMARY"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Get the last row of comparison data
        comparison_ws = self.workbook["July-August Comparison"]
        last_row = comparison_ws.max_row
        
        # Summary statistics
        summary_data = [
            ["Metric", "Value", "Formula", "Description"],
            ["Total Existing Users", f"=COUNTA('July-August Comparison'!A:A)-1", "=COUNTA('July-August Comparison'!A:A)-1", "Users in both July and August"],
            ["", "", "", ""],
            ["LOGIN COUNT STATISTICS", "", "", ""],
            ["Average Login Count Increase", f"=AVERAGE('July-August Comparison'!E2:E{last_row})", f"=AVERAGE('July-August Comparison'!E2:E{last_row})", "Average change in login count"],
            ["Median Login Count Increase", f"=MEDIAN('July-August Comparison'!E2:E{last_row})", f"=MEDIAN('July-August Comparison'!E2:E{last_row})", "Median change in login count"],
            ["Max Login Count Increase", f"=MAX('July-August Comparison'!E2:E{last_row})", f"=MAX('July-August Comparison'!E2:E{last_row})", "Maximum increase"],
            ["Min Login Count Increase", f"=MIN('July-August Comparison'!E2:E{last_row})", f"=MIN('July-August Comparison'!E2:E{last_row})", "Minimum increase"],
            ["Users with Positive Increase", f"=COUNTIF('July-August Comparison'!E2:E{last_row},\">0\")", f"=COUNTIF('July-August Comparison'!E2:E{last_row},\">0\")", "Users with more logins in August"],
            ["Users with Negative Increase", f"=COUNTIF('July-August Comparison'!E2:E{last_row},\"<0\")", f"=COUNTIF('July-August Comparison'!E2:E{last_row},\"<0\")", "Users with fewer logins in August"],
            ["", "", "", ""],
            ["AVERAGE LOGIN TIME STATISTICS", "", "", ""],
            ["Average Time Increase (seconds)", f"=AVERAGE('July-August Comparison'!I2:I{last_row})", f"=AVERAGE('July-August Comparison'!I2:I{last_row})", "Average change in login time"],
            ["Median Time Increase (seconds)", f"=MEDIAN('July-August Comparison'!I2:I{last_row})", f"=MEDIAN('July-August Comparison'!I2:I{last_row})", "Median change in login time"],
            ["Max Time Increase (seconds)", f"=MAX('July-August Comparison'!I2:I{last_row})", f"=MAX('July-August Comparison'!I2:I{last_row})", "Maximum time increase"],
            ["Min Time Increase (seconds)", f"=MIN('July-August Comparison'!I2:I{last_row})", f"=MIN('July-August Comparison'!I2:I{last_row})", "Minimum time increase"],
            ["Users with Positive Time Increase", f"=COUNTIF('July-August Comparison'!I2:I{last_row},\">0\")", f"=COUNTIF('July-August Comparison'!I2:I{last_row},\">0\")", "Users with longer login times"],
            ["Users with Negative Time Increase", f"=COUNTIF('July-August Comparison'!I2:I{last_row},\"<0\")", f"=COUNTIF('July-August Comparison'!I2:I{last_row},\"<0\")", "Users with shorter login times"],
            ["", "", "", ""],
            ["VWE STATISTICS", "", "", ""],
            ["Average VWE Increase", f"=AVERAGE('July-August Comparison'!M2:M{last_row})", f"=AVERAGE('July-August Comparison'!M2:M{last_row})", "Average change in VWE"],
            ["Median VWE Increase", f"=MEDIAN('July-August Comparison'!M2:M{last_row})", f"=MEDIAN('July-August Comparison'!M2:M{last_row})", "Median change in VWE"],
            ["Max VWE Increase", f"=MAX('July-August Comparison'!M2:M{last_row})", f"=MAX('July-August Comparison'!M2:M{last_row})", "Maximum VWE increase"],
            ["Min VWE Increase", f"=MIN('July-August Comparison'!M2:M{last_row})", f"=MIN('July-August Comparison'!M2:M{last_row})", "Minimum VWE increase"],
            ["Users with Positive VWE Increase", f"=COUNTIF('July-August Comparison'!M2:M{last_row},\">0\")", f"=COUNTIF('July-August Comparison'!M2:M{last_row},\">0\")", "Users with more VWE in August"],
            ["Users with Negative VWE Increase", f"=COUNTIF('July-August Comparison'!M2:M{last_row},\"<0\")", f"=COUNTIF('July-August Comparison'!M2:M{last_row},\"<0\")", "Users with less VWE in August"],
        ]
        
        # Add summary data
        for row_idx, row_data in enumerate(summary_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 2:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                elif col_idx == 1 and value in ["LOGIN COUNT STATISTICS", "AVERAGE LOGIN TIME STATISTICS", "VWE STATISTICS"]:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # Auto-adjust columns
        self.auto_adjust_columns(ws)
        
        print("Created Summary Statistics sheet")
    
    def create_formulas_guide_sheet(self):
        """Create a guide sheet explaining the formulas used"""
        ws = self.workbook.create_sheet("Formulas Guide")
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = "FORMULAS GUIDE - JULY-AUGUST COMPARISON"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Guide content
        guide_data = [
            ["Formula Type", "Example", "Purpose", "Explanation"],
            ["VLOOKUP", "=VLOOKUP(A2,'July Data'!B:C,2,FALSE)", "Find July Login Count", "Looks up email in July sheet, returns Login Count (column 2)"],
            ["VLOOKUP", "=VLOOKUP(A2,'August Data'!B:D,3,FALSE)", "Find August Avg Time", "Looks up email in August sheet, returns Avg Login Time (column 3)"],
            ["Simple Subtraction", "=D2-C2", "Calculate Increase", "Subtracts July value from August value"],
            ["Percentage Change", "=IF(C2=0,0,(D2-C2)/C2*100)", "Calculate % Change", "Calculates percentage change, handles division by zero"],
            ["AVERAGE", "=AVERAGE('July-August Comparison'!E2:E131)", "Average Increase", "Calculates average of all increases"],
            ["MEDIAN", "=MEDIAN('July-August Comparison'!E2:E131)", "Median Increase", "Calculates median of all increases"],
            ["MAX/MIN", "=MAX('July-August Comparison'!E2:E131)", "Max/Min Values", "Finds maximum or minimum values"],
            ["COUNTIF", "=COUNTIF('July-August Comparison'!E2:E131,\">0\")", "Count Positive", "Counts cells with positive values"],
            ["COUNTIF", "=COUNTIF('July-August Comparison'!E2:E131,\"<0\")", "Count Negative", "Counts cells with negative values"],
            ["", "", "", ""],
            ["COLUMN MAPPINGS", "", "", ""],
            ["July Data", "Column B = Email, Column C = Login Count, Column D = Avg Login Time, Column K = VWE", "", ""],
            ["August Data", "Column B = Email, Column C = Web Sessions, Column D = Avg Login Time, Column K = VWE", "", ""],
            ["", "", "", ""],
            ["USAGE INSTRUCTIONS", "", "", ""],
            ["1. Update Data", "Replace July Data and August Data sheets with new data", "", "The formulas will automatically recalculate"],
            ["2. View Results", "Check July-August Comparison sheet for individual user changes", "", "All calculations are automatic"],
            ["3. Summary", "Review Summary Statistics sheet for overall trends", "", "Statistics update automatically"],
            ["4. Customize", "Modify formulas in comparison sheet as needed", "", "Add new metrics or change calculations"],
        ]
        
        # Add guide data
        for row_idx, row_data in enumerate(guide_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 2:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                elif col_idx == 1 and value in ["COLUMN MAPPINGS", "USAGE INSTRUCTIONS"]:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # Auto-adjust columns
        self.auto_adjust_columns(ws)
        
        print("Created Formulas Guide sheet")
    
    def style_header(self, ws, row):
        """Style the header row"""
        for cell in ws[row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
    
    def auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def save_workbook(self, filename="July_August_Comparison_With_Formulas.xlsx"):
        """Save the workbook"""
        try:
            self.workbook.save(filename)
            print(f"Excel workbook saved as: {filename}")
            return True
        except Exception as e:
            print(f"Error saving workbook: {e}")
            return False

def main():
    print("CREATING EXCEL COMPARISON WORKBOOK WITH FORMULAS")
    print("="*60)
    
    # Initialize creator
    creator = ExcelComparisonCreator("August Export_SD 2 Sept.xlsx")
    
    # Load source data
    if not creator.load_source_data():
        return
    
    # Create workbook
    if not creator.create_workbook():
        return
    
    # Save workbook
    if creator.save_workbook():
        print("\n" + "="*60)
        print("EXCEL WORKBOOK CREATION COMPLETE")
        print("="*60)
        print("Created workbook: July_August_Comparison_With_Formulas.xlsx")
        print("\nThis workbook contains:")
        print("1. July Data sheet - Raw July data")
        print("2. August Data sheet - Raw August data")
        print("3. July-August Comparison sheet - Automatic calculations with formulas")
        print("4. Summary Statistics sheet - Overall statistics with formulas")
        print("5. Formulas Guide sheet - Explanation of all formulas used")
        print("\nAll calculations are automatic and will update when you change the data!")

if __name__ == "__main__":
    main()
