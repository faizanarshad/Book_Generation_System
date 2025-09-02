#!/usr/bin/env python3
"""
July-August User Activity Comparison
This script compares user activity between July and August tabs,
focusing on existing users (matching emails) and calculating increases
in Login Count, Avg Login Time, and VWE.
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import numpy as np

class JulyAugustComparison:
    def __init__(self, file_path):
        self.file_path = file_path
        self.july_data = None
        self.august_data = None
        self.comparison_results = None
        
    def load_data(self):
        """Load data from both July and August sheets"""
        try:
            # Load July data
            self.july_data = pd.read_excel(self.file_path, sheet_name='July ')
            print(f"July data loaded: {len(self.july_data)} rows")
            print(f"July columns: {list(self.july_data.columns)}")
            
            # Load August data
            self.august_data = pd.read_excel(self.file_path, sheet_name='August')
            print(f"August data loaded: {len(self.august_data)} rows")
            print(f"August columns: {list(self.august_data.columns)}")
            
            return True
            
        except Exception as e:
            print(f"Error loading data: {e}")
            return False
    
    def find_existing_users(self):
        """Find users that exist in both July and August (based on email)"""
        if self.july_data is None or self.august_data is None:
            print("Data not loaded. Please load data first.")
            return None
            
        try:
            # Get unique emails from both months
            july_emails = set(self.july_data['Email'].dropna().astype(str))
            august_emails = set(self.august_data['Email'].dropna().astype(str))
            
            # Find common emails (existing users)
            existing_emails = july_emails.intersection(august_emails)
            
            print(f"July unique users: {len(july_emails)}")
            print(f"August unique users: {len(august_emails)}")
            print(f"Existing users (in both months): {len(existing_emails)}")
            
            return existing_emails
            
        except Exception as e:
            print(f"Error finding existing users: {e}")
            return None
    
    def calculate_increases(self, existing_emails):
        """Calculate increases for existing users between July and August"""
        if existing_emails is None:
            print("No existing users found.")
            return None
            
        try:
            results = []
            
            for email in existing_emails:
                # Get July data for this user
                july_user = self.july_data[self.july_data['Email'].astype(str) == email]
                august_user = self.august_data[self.august_data['Email'].astype(str) == email]
                
                if len(july_user) > 0 and len(august_user) > 0:
                    # Get the first row for each user (in case of duplicates)
                    july_row = july_user.iloc[0]
                    august_row = august_user.iloc[0]
                    
                    # Extract values (handle NaN values)
                    july_login_count = july_row.get('Login Count', 0) if pd.notna(july_row.get('Login Count')) else 0
                    august_login_count = august_row.get('Login Count', 0) if pd.notna(august_row.get('Login Count')) else 0
                    
                    july_avg_time = july_row.get('Avg Login Time', 0) if pd.notna(july_row.get('Avg Login Time')) else 0
                    august_avg_time = august_row.get('Avg Login Time', 0) if pd.notna(august_row.get('Avg Login Time')) else 0
                    
                    july_vwe = july_row.get('Virtual Work Experience', 0) if pd.notna(july_row.get('Virtual Work Experience')) else 0
                    august_vwe = august_row.get('Virtual Work Experience', 0) if pd.notna(august_row.get('Virtual Work Experience')) else 0
                    
                    # Calculate increases
                    login_count_increase = august_login_count - july_login_count
                    avg_time_increase = august_avg_time - july_avg_time
                    vwe_increase = august_vwe - july_vwe
                    
                    # Calculate percentage increases (avoid division by zero)
                    login_count_pct = (login_count_increase / july_login_count * 100) if july_login_count > 0 else 0
                    avg_time_pct = (avg_time_increase / july_avg_time * 100) if july_avg_time > 0 else 0
                    vwe_pct = (vwe_increase / july_vwe * 100) if july_vwe > 0 else 0
                    
                    results.append({
                        'Email': email,
                        'First Name': august_row.get('First name', ''),
                        'July Login Count': july_login_count,
                        'August Login Count': august_login_count,
                        'Login Count Increase': login_count_increase,
                        'Login Count % Increase': round(login_count_pct, 2),
                        'July Avg Login Time': july_avg_time,
                        'August Avg Login Time': august_avg_time,
                        'Avg Time Increase (seconds)': avg_time_increase,
                        'Avg Time % Increase': round(avg_time_pct, 2),
                        'July VWE': july_vwe,
                        'August VWE': august_vwe,
                        'VWE Increase': vwe_increase,
                        'VWE % Increase': round(vwe_pct, 2)
                    })
            
            self.comparison_results = pd.DataFrame(results)
            return self.comparison_results
            
        except Exception as e:
            print(f"Error calculating increases: {e}")
            return None
    
    def generate_summary_statistics(self):
        """Generate summary statistics for the comparison"""
        if self.comparison_results is None:
            print("No comparison results available.")
            return None
            
        try:
            summary = {
                'Total Existing Users': len(self.comparison_results),
                'Login Count Increases': {
                    'Average Increase': round(self.comparison_results['Login Count Increase'].mean(), 2),
                    'Median Increase': round(self.comparison_results['Login Count Increase'].median(), 2),
                    'Max Increase': self.comparison_results['Login Count Increase'].max(),
                    'Min Increase': self.comparison_results['Login Count Increase'].min(),
                    'Users with Positive Increase': len(self.comparison_results[self.comparison_results['Login Count Increase'] > 0]),
                    'Users with Negative Increase': len(self.comparison_results[self.comparison_results['Login Count Increase'] < 0])
                },
                'Avg Login Time Increases': {
                    'Average Increase (seconds)': round(self.comparison_results['Avg Time Increase (seconds)'].mean(), 2),
                    'Median Increase (seconds)': round(self.comparison_results['Avg Time Increase (seconds)'].median(), 2),
                    'Max Increase (seconds)': self.comparison_results['Avg Time Increase (seconds)'].max(),
                    'Min Increase (seconds)': self.comparison_results['Avg Time Increase (seconds)'].min(),
                    'Users with Positive Increase': len(self.comparison_results[self.comparison_results['Avg Time Increase (seconds)'] > 0]),
                    'Users with Negative Increase': len(self.comparison_results[self.comparison_results['Avg Time Increase (seconds)'] < 0])
                },
                'VWE Increases': {
                    'Average Increase': round(self.comparison_results['VWE Increase'].mean(), 2),
                    'Median Increase': round(self.comparison_results['VWE Increase'].median(), 2),
                    'Max Increase': self.comparison_results['VWE Increase'].max(),
                    'Min Increase': self.comparison_results['VWE Increase'].min(),
                    'Users with Positive Increase': len(self.comparison_results[self.comparison_results['VWE Increase'] > 0]),
                    'Users with Negative Increase': len(self.comparison_results[self.comparison_results['VWE Increase'] < 0])
                }
            }
            
            return summary
            
        except Exception as e:
            print(f"Error generating summary: {e}")
            return None
    
    def save_results_to_excel(self, output_filename="July_August_Comparison_Results.xlsx"):
        """Save the comparison results to a new Excel file"""
        if self.comparison_results is None:
            print("No results to save.")
            return False
            
        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Comparison Results"
            
            # Write the comparison results
            for r, row in enumerate(self.comparison_results.itertuples(), 1):
                for c, value in enumerate(row[1:], 1):  # Skip index
                    ws.cell(row=r, column=c, value=value)
            
            # Add headers
            headers = list(self.comparison_results.columns)
            for c, header in enumerate(headers, 1):
                ws.cell(row=1, column=c, value=header)
                ws.cell(row=1, column=c).font = Font(bold=True)
                ws.cell(row=1, column=c).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save the file
            wb.save(output_filename)
            print(f"Results saved to: {output_filename}")
            return True
            
        except Exception as e:
            print(f"Error saving results: {e}")
            return False
    
    def print_detailed_results(self, top_n=10):
        """Print detailed results for top performers"""
        if self.comparison_results is None:
            print("No results to display.")
            return
            
        print("\n" + "="*100)
        print("DETAILED COMPARISON RESULTS")
        print("="*100)
        
        # Top performers by Login Count increase
        print(f"\nTOP {top_n} USERS BY LOGIN COUNT INCREASE:")
        print("-" * 80)
        top_login = self.comparison_results.nlargest(top_n, 'Login Count Increase')
        for _, row in top_login.iterrows():
            print(f"{row['First Name']} ({row['Email'][:30]}...): "
                  f"July: {row['July Login Count']} → August: {row['August Login Count']} "
                  f"(+{row['Login Count Increase']}, {row['Login Count % Increase']}%)")
        
        # Top performers by Avg Login Time increase
        print(f"\nTOP {top_n} USERS BY AVERAGE LOGIN TIME INCREASE:")
        print("-" * 80)
        top_time = self.comparison_results.nlargest(top_n, 'Avg Time Increase (seconds)')
        for _, row in top_time.iterrows():
            print(f"{row['First Name']} ({row['Email'][:30]}...): "
                  f"July: {row['July Avg Login Time']}s → August: {row['August Avg Login Time']}s "
                  f"(+{row['Avg Time Increase (seconds)']}s, {row['Avg Time % Increase']}%)")
        
        # Top performers by VWE increase
        print(f"\nTOP {top_n} USERS BY VWE INCREASE:")
        print("-" * 80)
        top_vwe = self.comparison_results.nlargest(top_n, 'VWE Increase')
        for _, row in top_vwe.iterrows():
            print(f"{row['First Name']} ({row['Email'][:30]}...): "
                  f"July: {row['July VWE']} → August: {row['August VWE']} "
                  f"(+{row['VWE Increase']}, {row['VWE % Increase']}%)")

def main():
    # Initialize the comparison
    comparison = JulyAugustComparison("August Export_SD 2 Sept.xlsx")
    
    print("JULY-AUGUST USER ACTIVITY COMPARISON")
    print("="*50)
    
    # Load data
    if not comparison.load_data():
        return
    
    # Find existing users
    existing_users = comparison.find_existing_users()
    if existing_users is None:
        return
    
    # Calculate increases
    results = comparison.calculate_increases(existing_users)
    if results is None:
        return
    
    # Generate and display summary statistics
    summary = comparison.generate_summary_statistics()
    if summary:
        print("\n" + "="*60)
        print("SUMMARY STATISTICS")
        print("="*60)
        print(f"Total Existing Users: {summary['Total Existing Users']}")
        
        print(f"\nLOGIN COUNT INCREASES:")
        print(f"  Average Increase: {summary['Login Count Increases']['Average Increase']}")
        print(f"  Median Increase: {summary['Login Count Increases']['Median Increase']}")
        print(f"  Max Increase: {summary['Login Count Increases']['Max Increase']}")
        print(f"  Min Increase: {summary['Login Count Increases']['Min Increase']}")
        print(f"  Users with Positive Increase: {summary['Login Count Increases']['Users with Positive Increase']}")
        print(f"  Users with Negative Increase: {summary['Login Count Increases']['Users with Negative Increase']}")
        
        print(f"\nAVERAGE LOGIN TIME INCREASES:")
        print(f"  Average Increase: {summary['Avg Login Time Increases']['Average Increase (seconds)']} seconds")
        print(f"  Median Increase: {summary['Avg Login Time Increases']['Median Increase (seconds)']} seconds")
        print(f"  Max Increase: {summary['Avg Login Time Increases']['Max Increase (seconds)']} seconds")
        print(f"  Min Increase: {summary['Avg Login Time Increases']['Min Increase (seconds)']} seconds")
        print(f"  Users with Positive Increase: {summary['Avg Login Time Increases']['Users with Positive Increase']}")
        print(f"  Users with Negative Increase: {summary['Avg Login Time Increases']['Users with Negative Increase']}")
        
        print(f"\nVWE INCREASES:")
        print(f"  Average Increase: {summary['VWE Increases']['Average Increase']}")
        print(f"  Median Increase: {summary['VWE Increases']['Median Increase']}")
        print(f"  Max Increase: {summary['VWE Increases']['Max Increase']}")
        print(f"  Min Increase: {summary['VWE Increases']['Min Increase']}")
        print(f"  Users with Positive Increase: {summary['VWE Increases']['Users with Positive Increase']}")
        print(f"  Users with Negative Increase: {summary['VWE Increases']['Users with Negative Increase']}")
    
    # Print detailed results
    comparison.print_detailed_results(top_n=10)
    
    # Save results to Excel
    comparison.save_results_to_excel()
    
    print("\n" + "="*60)
    print("ANALYSIS COMPLETE")
    print("="*60)
    print("Detailed results have been saved to 'July_August_Comparison_Results.xlsx'")
    print("This file contains all user comparisons with increases in Login Count, Avg Login Time, and VWE.")

if __name__ == "__main__":
    main()
