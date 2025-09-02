import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def analyze_existing_users():
    """Analyze existing users from July to August and calculate activity increases"""
    
    try:
        # Read the Excel file
        file_path = "August Export_SD 2 Sept_modified.xlsx"
        print(f"Reading data from: {file_path}")
        
        # Read both sheets
        july_df = pd.read_excel(file_path, sheet_name='July ')
        august_df = pd.read_excel(file_path, sheet_name='August')
        
        print(f"July data shape: {july_df.shape}")
        print(f"August data shape: {august_df.shape}")
        
        # Clean column names (remove extra spaces)
        july_df.columns = july_df.columns.str.strip()
        august_df.columns = august_df.columns.str.strip()
        
        print(f"\nJuly columns: {list(july_df.columns)}")
        print(f"August columns: {list(august_df.columns)}")
        
        # Find existing users (emails that appear in both sheets)
        july_emails = set(july_df['Email'].str.lower())
        august_emails = set(august_df['Email'].str.lower())
        existing_emails = july_emails.intersection(august_emails)
        
        print(f"\nTotal unique emails in July: {len(july_emails)}")
        print(f"Total unique emails in August: {len(august_emails)}")
        print(f"Existing users (emails in both): {len(existing_emails)}")
        
        # Create comparison dataframe for existing users
        comparison_data = []
        
        for email in existing_emails:
            # Get July data
            july_row = july_df[july_df['Email'].str.lower() == email].iloc[0]
            # Get August data
            august_row = august_df[august_df['Email'].str.lower() == email].iloc[0]
            
            # Extract values for comparison
            july_login_count = july_row['Login Count']
            august_web_sessions = august_row['Web sessions']
            
            july_avg_time = july_row['Avg Login Time']
            august_avg_time = august_row['Avg Login Time']
            
            july_vwe = july_row['Virtual Work Experience']
            august_vwe = august_row['Virtual Work Experience']
            
            # Calculate increases
            login_increase = august_web_sessions - july_login_count
            time_increase = august_avg_time - july_avg_time
            vwe_increase = august_vwe - july_vwe if pd.notna(august_vwe) and pd.notna(july_vwe) else np.nan
            
            comparison_data.append({
                'Email': email,
                'First Name (July)': july_row['First name'],
                'First Name (August)': august_row['First name'],
                'July_Login_Count': july_login_count,
                'August_Web_Sessions': august_web_sessions,
                'Login_Increase': login_increase,
                'July_Avg_Login_Time': july_avg_time,
                'August_Avg_Login_Time': august_avg_time,
                'Time_Increase_Seconds': time_increase,
                'July_VWE': july_vwe,
                'August_VWE': august_vwe,
                'VWE_Increase': vwe_increase
            })
        
        # Create comparison dataframe
        comparison_df = pd.DataFrame(comparison_data)
        
        print(f"\nComparison data shape: {comparison_df.shape}")
        print("\nSample comparison data:")
        print(comparison_df.head())
        
        # Calculate summary statistics
        print("\n" + "="*60)
        print("SUMMARY STATISTICS FOR EXISTING USERS")
        print("="*60)
        
        # Login/Web Sessions Analysis
        print(f"\nLOGIN COUNT / WEB SESSIONS ANALYSIS:")
        print(f"Average July Login Count: {comparison_df['July_Login_Count'].mean():.2f}")
        print(f"Average August Web Sessions: {comparison_df['August_Web_Sessions'].mean():.2f}")
        print(f"Average Increase: {comparison_df['Login_Increase'].mean():.2f}")
        print(f"Users with Increased Activity: {(comparison_df['Login_Increase'] > 0).sum()}")
        print(f"Users with Decreased Activity: {(comparison_df['Login_Increase'] < 0).sum()}")
        print(f"Users with Same Activity: {(comparison_df['Login_Increase'] == 0).sum()}")
        
        # Time Analysis
        print(f"\nAVERAGE LOGIN TIME ANALYSIS (seconds):")
        print(f"Average July Time: {comparison_df['July_Avg_Login_Time'].mean():.2f}")
        print(f"Average August Time: {comparison_df['August_Avg_Login_Time'].mean():.2f}")
        print(f"Average Time Increase: {comparison_df['Time_Increase_Seconds'].mean():.2f}")
        print(f"Users with Increased Time: {(comparison_df['Time_Increase_Seconds'] > 0).sum()}")
        print(f"Users with Decreased Time: {(comparison_df['Time_Increase_Seconds'] < 0).sum()}")
        
        # VWE Analysis
        print(f"\nVIRTUAL WORK EXPERIENCE ANALYSIS:")
        vwe_comparison = comparison_df.dropna(subset=['VWE_Increase'])
        if not vwe_comparison.empty:
            print(f"Users with VWE data in both months: {len(vwe_comparison)}")
            print(f"Average July VWE: {vwe_comparison['July_VWE'].mean():.2f}")
            print(f"Average August VWE: {vwe_comparison['August_VWE'].mean():.2f}")
            print(f"Average VWE Increase: {vwe_comparison['VWE_Increase'].mean():.2f}")
            print(f"Users with Increased VWE: {(vwe_comparison['VWE_Increase'] > 0).sum()}")
            print(f"Users with Decreased VWE: {(vwe_comparison['VWE_Increase'] < 0).sum()}")
        
        # Save comparison to new Excel file with formulas
        output_file = "August_Export_SD_2_Sept_updated.xlsx"
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write comparison data
            comparison_df.to_excel(writer, sheet_name='July_August_Comparison', index=False)
            
            # Copy original sheets
            for sheet_name in ['July ', 'August', 'Sheet7']:
                sheet_df = pd.read_excel(file_path, sheet_name=sheet_name)
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"\nComparison data saved to: {output_file}")
        
        # Create a summary sheet with formulas
        create_summary_with_formulas(output_file, comparison_df)
        
        return comparison_df
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return None

def create_summary_with_formulas(output_file, comparison_df):
    """Create a summary sheet with Excel formulas for the analysis"""
    
    try:
        # Load the workbook
        workbook = load_workbook(output_file)
        
        # Create summary sheet
        if 'Summary_With_Formulas' in workbook.sheetnames:
            workbook.remove(workbook['Summary_With_Formulas'])
        
        summary_sheet = workbook.create_sheet('Summary_With_Formulas')
        
        # Add headers
        headers = [
            'Metric', 'Value', 'Formula Used', 'Description'
        ]
        
        for col, header in enumerate(headers, 1):
            summary_sheet.cell(row=1, column=col, value=header)
        
        # Add data with formulas
        row = 2
        
        # Total existing users
        summary_sheet.cell(row=row, column=1, value='Total Existing Users')
        summary_sheet.cell(row=row, column=2, value=len(comparison_df))
        summary_sheet.cell(row=row, column=3, value=f'=COUNTA(July_August_Comparison!A:A)-1')
        summary_sheet.cell(row=row, column=4, value='Count of users present in both July and August')
        row += 1
        
        # Login increase stats
        summary_sheet.cell(row=row, column=1, value='Average Login Increase')
        summary_sheet.cell(row=row, column=2, value=comparison_df['Login_Increase'].mean())
        summary_sheet.cell(row=row, column=3, value=f'=AVERAGE(July_August_Comparison!F:F)')
        summary_sheet.cell(row=row, column=4, value='Average increase in web sessions from July to August')
        row += 1
        
        summary_sheet.cell(row=row, column=1, value='Users with Increased Login Activity')
        summary_sheet.cell(row=row, column=2, value=(comparison_df['Login_Increase'] > 0).sum())
        summary_sheet.cell(row=row, column=3, value=f'=COUNTIF(July_August_Comparison!F:F,">0")')
        summary_sheet.cell(row=row, column=4, value='Number of users with more web sessions in August')
        row += 1
        
        # Time increase stats
        summary_sheet.cell(row=row, column=1, value='Average Time Increase (seconds)')
        summary_sheet.cell(row=row, column=2, value=comparison_df['Time_Increase_Seconds'].mean())
        summary_sheet.cell(row=row, column=3, value=f'=AVERAGE(July_August_Comparison!H:H)')
        summary_sheet.cell(row=row, column=4, value='Average increase in login time from July to August')
        row += 1
        
        summary_sheet.cell(row=row, column=1, value='Users with Increased Login Time')
        summary_sheet.cell(row=row, column=2, value=(comparison_df['Time_Increase_Seconds'] > 0).sum())
        summary_sheet.cell(row=row, column=3, value=f'=COUNTIF(July_August_Comparison!H:H,">0")')
        summary_sheet.cell(row=row, column=4, value='Number of users with longer login times in August')
        row += 1
        
        # VWE stats
        vwe_comparison = comparison_df.dropna(subset=['VWE_Increase'])
        if not vwe_comparison.empty:
            summary_sheet.cell(row=row, column=1, value='Users with VWE Data in Both Months')
            summary_sheet.cell(row=row, column=2, value=len(vwe_comparison))
            summary_sheet.cell(row=row, column=3, value=f'=COUNTA(July_August_Comparison!K:K)')
            summary_sheet.cell(row=row, column=4, value='Count of users with VWE data in both months')
            row += 1
            
            summary_sheet.cell(row=row, column=1, value='Average VWE Increase')
            summary_sheet.cell(row=row, column=2, value=vwe_comparison['VWE_Increase'].mean())
            summary_sheet.cell(row=row, column=3, value=f'=AVERAGE(July_August_Comparison!K:K)')
            summary_sheet.cell(row=row, column=4, value='Average increase in VWE from July to August')
            row += 1
        
        # Add formula examples
        row += 1
        summary_sheet.cell(row=row, column=1, value='FORMULA EXAMPLES')
        summary_sheet.cell(row=row, column=2, value='')
        summary_sheet.cell(row=row, column=3, value='')
        summary_sheet.cell(row=row, column=4, value='')
        row += 1
        
        summary_sheet.cell(row=row, column=1, value='Login Increase Formula')
        summary_sheet.cell(row=row, column=2, value='')
        summary_sheet.cell(row=row, column=3, value='=August_Web_Sessions - July_Login_Count')
        summary_sheet.cell(row=row, column=4, value='Calculate increase in activity')
        row += 1
        
        summary_sheet.cell(row=row, column=1, value='Time Increase Formula')
        summary_sheet.cell(row=row, column=2, value='')
        summary_sheet.cell(row=row, column=3, value='=August_Avg_Login_Time - July_Avg_Login_Time')
        summary_sheet.cell(row=row, column=4, value='Calculate increase in login time')
        row += 1
        
        summary_sheet.cell(row=row, column=1, value='VWE Increase Formula')
        summary_sheet.cell(row=row, column=2, value='')
        summary_sheet.cell(row=row, column=3, value='=August_VWE - July_VWE')
        summary_sheet.cell(row=row, column=4, value='Calculate increase in VWE')
        
        # Save the workbook
        workbook.save(output_file)
        print(f"Summary sheet with formulas added to: {output_file}")
        
    except Exception as e:
        print(f"Error creating summary sheet: {e}")

if __name__ == "__main__":
    comparison_df = analyze_existing_users()
    if comparison_df is not None:
        print(f"\nAnalysis completed successfully!")
        print(f"Found {len(comparison_df)} existing users to compare")
