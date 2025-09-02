#!/usr/bin/env python3
"""
Add Industry Preferences Table with Excel Formulas to Master Excel File
This script adds the industry preferences table as a new sheet with Excel formulas
that automatically calculate the data from the August sheet.
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np
import re

def add_industry_preferences_with_formulas():
    """Add industry preferences table with Excel formulas to the master Excel file"""
    
    master_file = "August Export_SD 2 Sept.xlsx"
    
    # Load the master workbook
    print(f"Loading master Excel file: {master_file}")
    wb = load_workbook(master_file)
    
    # Check existing sheets
    print(f"Existing sheets: {wb.sheetnames}")
    
    # Load data from existing sheets to understand structure
    print("Analyzing data structure...")
    august_data = pd.read_excel(master_file, sheet_name='August')
    industry_mapping = {}
    
    # Load industry mapping from Sheet 7
    sheet7_data = pd.read_excel(master_file, sheet_name='Sheet7', header=None)
    for i in range(2, len(sheet7_data)):
        row = sheet7_data.iloc[i]
        if pd.notna(row[1]) and pd.notna(row[2]):
            industry_number = int(row[1])
            industry_name = str(row[2]).strip()
            industry_mapping[industry_number] = industry_name
    
    print(f"Loaded {len(industry_mapping)} industry mappings")
    
    # Define the exact industry order requested
    industry_order = [
        'Accounting', 'Advertising, Media, Journalism, and Communications',
        'Agriculture and Environment', 'Animals and Vet', 'Architecture',
        'Arts, Humanities, and Politics', 'Building and Construction',
        'Business and Commerce', 'Community and Social Work',
        'Creative Arts and Music', 'Design', 'Economics and Finance',
        'Education, Childcare and Teaching', 'Engineering', 'Entrepreneur',
        'Food and Beverage', 'Government, Defence and Policing',
        'Hair and Beauty', 'Health and Sport Sciences', 'Law',
        'Marketing and Public Relations', 'Mathematics',
        'Medical Sciences and Medicine', 'Nursing and Midwifery',
        'Property and Real Estate', 'Psychology', 'Science', 'Technology',
        'Trades and Mining', 'Sports', 'Transport, Tourism and Hospitality',
        'Fashion', 'Australian Defence Force', 'Energy'
    ]
    
    # Create new sheet in master workbook
    print("Creating Industry Preferences with Formulas sheet in master workbook...")
    
    # Remove existing sheet if it exists
    if "Industry Preferences with Formulas" in wb.sheetnames:
        wb.remove(wb["Industry Preferences with Formulas"])
        print("Removed existing Industry Preferences with Formulas sheet")
    
    # Create new sheet
    ws = wb.create_sheet("Industry Preferences with Formulas")
    
    # Title
    ws.merge_cells('A1:L1')
    ws['A1'] = "INDUSTRY PREFERENCES BY FACULTY AND YEAR GROUP (WITH FORMULAS)"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Faculty headers
    ws.merge_cells('B2:F2')
    ws['B2'] = "Faculty of Engineering"
    ws.merge_cells('G2:K2')
    ws['G2'] = "Faculty of Arts and Social Sciences"
    
    # Style faculty headers
    for cell in ['B2', 'G2']:
        ws[cell].font = Font(bold=True, size=12)
        ws[cell].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        ws[cell].alignment = Alignment(horizontal='center')
    
    # Year headers
    year_headers = ['1st Year', '2nd Year', '3rd Year', '4th Year', '5th Year']
    for i, year in enumerate(year_headers):
        # Engineering years
        ws.cell(row=3, column=i+2, value=year)
        # Arts years
        ws.cell(row=3, column=i+7, value=year)
    
    # Style year headers
    for col in range(2, 12):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Industry column header
    ws['A3'] = "Industry"
    ws['A3'].font = Font(bold=True)
    ws['A3'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Add industry names
    for row_idx, industry in enumerate(industry_order, 4):
        ws.cell(row=row_idx, column=1, value=industry)
        ws.cell(row=row_idx, column=1).font = Font(bold=True)
    
    # Add formulas for Engineering faculty
    print("Adding Engineering faculty formulas...")
    for row_idx, industry in enumerate(industry_order, 4):
        # Find the industry number for this industry
        industry_num = None
        for num, name in industry_mapping.items():
            if name == industry:
                industry_num = num
                break
        
        if industry_num is not None:
            # Add formulas for each year
            for col_idx, year in enumerate(year_headers):
                # Formula to count students with this industry preference in Engineering faculty
                # This formula will automatically update when August sheet data changes
                formula = f'=COUNTIFS(\'August\'!J:J,"*{industry_num}*",\'August\'!I:I,"*Faculty of Engineering*",\'August\'!P:P,"*{year}*")'
                ws.cell(row=row_idx, column=col_idx+2, value=formula)
    
    # Add formulas for Arts faculty
    print("Adding Arts faculty formulas...")
    for row_idx, industry in enumerate(industry_order, 4):
        # Find the industry number for this industry
        industry_num = None
        for num, name in industry_mapping.items():
            if name == industry:
                industry_num = num
                break
        
        if industry_num is not None:
            # Add formulas for each year
            for col_idx, year in enumerate(year_headers):
                # Formula to count students with this industry preference in Arts faculty
                formula = f'=COUNTIFS(\'August\'!J:J,"*{industry_num}*",\'August\'!I:I,"*Faculty of Arts and Social Sciences*",\'August\'!P:P,"*{year}*")'
                ws.cell(row=row_idx, column=col_idx+7, value=formula)
    
    # Add summary formulas at the bottom
    print("Adding summary formulas...")
    
    # Add totals row
    total_row = len(industry_order) + 4
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=1).fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # Engineering totals
    for col_idx in range(2, 7):
        col_letter = get_column_letter(col_idx)
        formula = f'=SUM({col_letter}4:{col_letter}{total_row-1})'
        ws.cell(row=total_row, column=col_idx, value=formula)
        ws.cell(row=total_row, column=col_idx).font = Font(bold=True)
        ws.cell(row=total_row, column=col_idx).fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # Arts totals
    for col_idx in range(7, 12):
        col_letter = get_column_letter(col_idx)
        formula = f'=SUM({col_letter}4:{col_letter}{total_row-1})'
        ws.cell(row=total_row, column=col_idx, value=formula)
        ws.cell(row=total_row, column=col_idx).font = Font(bold=True)
        ws.cell(row=total_row, column=col_idx).fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # Grand total
    ws.cell(row=total_row, column=12, value=f'=SUM(B{total_row}:K{total_row})')
    ws.cell(row=total_row, column=12).font = Font(bold=True)
    ws.cell(row=total_row, column=12).fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create a formula explanation sheet
    print("Creating formula explanation sheet...")
    if "Formula Guide" in wb.sheetnames:
        wb.remove(wb["Formula Guide"])
    
    ws_guide = wb.create_sheet("Formula Guide")
    
    # Title
    ws_guide.merge_cells('A1:C1')
    ws_guide['A1'] = "FORMULA EXPLANATION FOR INDUSTRY PREFERENCES TABLE"
    ws_guide['A1'].font = Font(size=16, bold=True)
    ws_guide['A1'].alignment = Alignment(horizontal='center')
    
    # Formula explanations
    guide_data = [
        ["Formula Type", "Example", "Explanation"],
        ["Industry Count", "=COUNTIFS('August'!J:J,\"*14*\",'August'!I:I,\"*Faculty of Engineering*\",'August'!P:P,\"*1st Year*\")", "Counts students in Engineering faculty, 1st year who prefer Engineering (industry 14)"],
        ["", "", ""],
        ["Formula Components:", "", ""],
        ["Column J (Industries)", "'August'!J:J", "Looks for industry numbers in column J"],
        ["Faculty Filter", "'August'!I:I,\"*Faculty of Engineering*\"", "Filters for Engineering faculty students"],
        ["Year Filter", "'August'!P:P,\"*1st Year*\"", "Filters for specific year group"],
        ["", "", ""],
        ["How to Use:", "", ""],
        ["1. Update Data", "Change data in August sheet", "Formulas automatically recalculate"],
        ["2. Add New Students", "Add rows to August sheet", "Formulas automatically include new data"],
        ["3. Modify Industries", "Change industry numbers in column J", "Formulas automatically update counts"],
        ["", "", ""],
        ["Column Mappings:", "", ""],
        ["Column I", "Faculty", "Contains faculty information"],
        ["Column J", "Industries", "Contains industry preference numbers"],
        ["Column P", "Course Year", "Contains year group information"],
        ["", "", ""],
        ["Industry Numbers:", "", ""],
        ["14", "Engineering", "From Sheet7 mapping"],
        ["28", "Technology", "From Sheet7 mapping"],
        ["27", "Science", "From Sheet7 mapping"],
        ["...", "...", "See Sheet7 for complete list"]
    ]
    
    for row_idx, row_data in enumerate(guide_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_guide.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            elif col_idx == 1 and value in ["Formula Components:", "How to Use:", "Column Mappings:", "Industry Numbers:"]:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    # Auto-adjust columns for guide sheet
    for column in ws_guide.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_guide.column_dimensions[column_letter].width = adjusted_width
    
    # Save the updated master workbook
    print("Saving updated master workbook...")
    wb.save(master_file)
    
    print(f"\n✅ Successfully added 'Industry Preferences with Formulas' sheet to {master_file}")
    print("The table now contains Excel formulas that automatically calculate from your data!")
    print("Also added 'Formula Guide' sheet to explain how the formulas work.")
    
    return True

if __name__ == "__main__":
    print("ADDING INDUSTRY PREFERENCES TABLE WITH EXCEL FORMULAS TO MASTER EXCEL FILE")
    print("="*70)
    
    try:
        success = add_industry_preferences_with_formulas()
        
        if success:
            print("\n" + "="*60)
            print("OPERATION COMPLETE")
            print("="*60)
            print("Your master Excel file now contains:")
            print("1. July sheet - July data")
            print("2. August sheet - August data")
            print("3. Sheet7 - Industry mappings")
            print("4. Industry Preferences with Formulas - NEW sheet with Excel formulas")
            print("5. Formula Guide - NEW sheet explaining how formulas work")
            print("\nKey Benefits:")
            print("✅ Formulas automatically calculate from August sheet data")
            print("✅ Updates automatically when you change the data")
            print("✅ No need to re-run scripts - just update the data!")
            print("✅ Professional Excel table with dynamic calculations")
            print("\nOpen the Excel file to see the new sheets with formulas!")
        
    except Exception as e:
        print(f"Error: {e}")
        print("Please make sure the master Excel file is not open in Excel when running this script.")
