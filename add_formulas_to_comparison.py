import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

def add_formulas_to_comparison():
    """Add Excel formulas to the comparison sheet for calculating increases"""
    
    try:
        file_path = "August_Export_SD_2_Sept_updated.xlsx"
        print(f"Adding formulas to: {file_path}")
        
        # Load the workbook
        workbook = load_workbook(file_path)
        
        # Get the comparison sheet
        if 'July_August_Comparison' not in workbook.sheetnames:
            print("Error: July_August_Comparison sheet not found!")
            return
        
        comparison_sheet = workbook['July_August_Comparison']
        
        # Add formula headers
        print("Adding formula headers...")
        
        # Find the last column and add formula columns
        max_col = comparison_sheet.max_column
        
        # Add formula column headers
        comparison_sheet.cell(row=1, column=max_col + 1, value="Login_Increase_Formula")
        comparison_sheet.cell(row=1, column=max_col + 2, value="Time_Increase_Formula")
        comparison_sheet.cell(row=1, column=max_col + 3, value="VWE_Increase_Formula")
        
        # Style the headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col in range(max_col + 1, max_col + 4):
            cell = comparison_sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        
        # Add formulas to each row
        print("Adding formulas to data rows...")
        
        for row in range(2, comparison_sheet.max_row + 1):
            # Login Increase Formula: =August_Web_Sessions - July_Login_Count
            login_formula = f"={comparison_sheet.cell(row=row, column=5).coordinate}-{comparison_sheet.cell(row=row, column=4).coordinate}"
            comparison_sheet.cell(row=row, column=max_col + 1, value=login_formula)
            
            # Time Increase Formula: =August_Avg_Login_Time - July_Avg_Login_Time
            time_formula = f"={comparison_sheet.cell(row=row, column=8).coordinate}-{comparison_sheet.cell(row=row, column=7).coordinate}"
            comparison_sheet.cell(row=row, column=max_col + 2, value=time_formula)
            
            # VWE Increase Formula: =August_VWE - July_VWE
            vwe_formula = f"={comparison_sheet.cell(row=row, column=11).coordinate}-{comparison_sheet.cell(row=row, column=10).coordinate}"
            comparison_sheet.cell(row=row, column=max_col + 3, value=vwe_formula)
        
        # Add a summary section at the bottom
        print("Adding summary section...")
        
        summary_start_row = comparison_sheet.max_row + 3
        
        # Summary headers
        summary_headers = [
            "SUMMARY STATISTICS",
            "Total Existing Users",
            "Average Login Increase",
            "Average Time Increase (seconds)",
            "Average VWE Increase",
            "Users with Increased Login Activity",
            "Users with Increased Login Time",
            "Users with Increased VWE"
        ]
        
        for i, header in enumerate(summary_headers):
            cell = comparison_sheet.cell(row=summary_start_row + i, column=1, value=header)
            cell.font = Font(bold=True)
            if i == 0:  # Main header
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Add summary formulas
        # Total users
        total_users_formula = f"=COUNTA(A:A)-1"
        comparison_sheet.cell(row=summary_start_row + 1, column=2, value=total_users_formula)
        
        # Average login increase
        avg_login_formula = f"=AVERAGE(F:F)"
        comparison_sheet.cell(row=summary_start_row + 2, column=2, value=avg_login_formula)
        
        # Average time increase
        avg_time_formula = f"=AVERAGE(H:H)"
        comparison_sheet.cell(row=summary_start_row + 3, column=2, value=avg_time_formula)
        
        # Average VWE increase
        avg_vwe_formula = f"=AVERAGE(K:K)"
        comparison_sheet.cell(row=summary_start_row + 4, column=2, value=avg_vwe_formula)
        
        # Count users with increased login activity
        increased_login_formula = f"=COUNTIF(F:F,\">0\")"
        comparison_sheet.cell(row=summary_start_row + 5, column=2, value=increased_login_formula)
        
        # Count users with increased login time
        increased_time_formula = f"=COUNTIF(H:H,\">0\")"
        comparison_sheet.cell(row=summary_start_row + 6, column=2, value=increased_time_formula)
        
        # Count users with increased VWE
        increased_vwe_formula = f"=COUNTIF(K:K,\">0\")"
        comparison_sheet.cell(row=summary_start_row + 7, column=2, value=increased_vwe_formula)
        
        # Add formula explanations
        formula_explanations = [
            "FORMULA EXPLANATIONS",
            "Login Increase: =August_Web_Sessions - July_Login_Count",
            "Time Increase: =August_Avg_Login_Time - July_Avg_Login_Time",
            "VWE Increase: =August_VWE - July_VWE",
            "Positive values = Increased activity in August",
            "Negative values = Decreased activity in August",
            "Zero = No change in activity"
        ]
        
        explanation_start_row = summary_start_row + 10
        
        for i, explanation in enumerate(formula_explanations):
            cell = comparison_sheet.cell(row=explanation_start_row + i, column=1, value=explanation)
            if i == 0:  # Main header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif i > 0 and i <= 3:  # Formula rows
                cell.font = Font(bold=True, color="0000FF")
        
        # Save the workbook
        workbook.save(file_path)
        print(f"Formulas and summary added successfully to: {file_path}")
        
        # Show what was added
        print(f"\nAdded to comparison sheet:")
        print(f"- 3 new formula columns (Login, Time, VWE increases)")
        print(f"- Summary statistics section with formulas")
        print(f"- Formula explanations")
        print(f"- Styled headers and sections")
        
    except Exception as e:
        print(f"Error adding formulas: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    add_formulas_to_comparison()
