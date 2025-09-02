#!/usr/bin/env python3
"""
Industry Preferences Analysis
This script analyzes student industry preferences by mapping numbers from column J
to industries in Sheet 7, and creates a table showing which industries are popular
with which students by faculty and year group.
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import numpy as np
import re

class IndustryPreferencesAnalysis:
    def __init__(self, file_path):
        self.file_path = file_path
        self.industry_mapping = {}
        self.august_data = None
        self.analysis_results = {}
        
    def load_industry_mapping(self):
        """Load industry number to name mapping from Sheet 7"""
        try:
            # Read Sheet 7
            df = pd.read_excel(self.file_path, sheet_name='Sheet7', header=None)
            
            # Extract industry mappings (rows 2 onwards)
            for i in range(2, len(df)):
                row = df.iloc[i]
                if pd.notna(row[1]) and pd.notna(row[2]):  # Check if both number and name exist
                    industry_number = int(row[1])
                    industry_name = str(row[2]).strip()
                    self.industry_mapping[industry_number] = industry_name
            
            print(f"Loaded {len(self.industry_mapping)} industry mappings:")
            for num, name in sorted(self.industry_mapping.items()):
                print(f"  {num}: {name}")
            
            return True
            
        except Exception as e:
            print(f"Error loading industry mapping: {e}")
            return False
    
    def load_august_data(self):
        """Load August data"""
        try:
            self.august_data = pd.read_excel(self.file_path, sheet_name='August')
            print(f"August data loaded: {len(self.august_data)} rows")
            return True
        except Exception as e:
            print(f"Error loading August data: {e}")
            return False
    
    def parse_industry_numbers(self, industry_string):
        """Parse industry numbers from string like '|14|15|12|27|28|'"""
        if pd.isna(industry_string):
            return []
        
        # Remove pipes and split by |
        numbers_str = str(industry_string).replace("'", "").replace("|", " ")
        numbers = re.findall(r'\d+', numbers_str)
        return [int(num) for num in numbers if int(num) in self.industry_mapping]
    
    def clean_faculty_name(self, faculty_string):
        """Clean faculty name by removing pipes and standardizing"""
        if pd.isna(faculty_string):
            return "Unknown"
        
        faculty = str(faculty_string).replace("'", "").replace("|", "").strip()
        
        # Standardize faculty names
        if "Faculty of Engineering" in faculty:
            return "Faculty of Engineering"
        elif "Faculty of Arts and Social Sciences" in faculty:
            return "Faculty of Arts and Social Sciences"
        elif "University of Sydney Business School" in faculty:
            return "University of Sydney Business School"
        elif "Faculty of Medicine and Health" in faculty:
            return "Faculty of Medicine and Health"
        elif "Sydney School of Architecture, Design and Planning" in faculty:
            return "Sydney School of Architecture, Design and Planning"
        elif "Sydney Law School" in faculty:
            return "Sydney Law School"
        elif "Sydney Conservatorium of Music" in faculty:
            return "Sydney Conservatorium of Music"
        else:
            return "Other"
    
    def clean_year_name(self, year_string):
        """Clean year name by removing pipes and standardizing"""
        if pd.isna(year_string):
            return "Unknown"
        
        year = str(year_string).replace("'", "").replace("|", "").strip()
        
        # Standardize year names
        if "1st Year" in year or year == "1":
            return "1st Year"
        elif "2nd Year" in year or year == "2":
            return "2nd Year"
        elif "3rd Year" in year or year == "3":
            return "3rd Year"
        elif "4th Year" in year or year == "4":
            return "4th Year"
        elif "5th Year" in year or year == "5":
            return "5th Year"
        else:
            return "Unknown"
    
    def create_industry_preferences_table(self):
        """Create the main industry preferences table"""
        if self.august_data is None:
            print("No August data loaded.")
            return None
            
        try:
            # Prepare data
            data = self.august_data.copy()
            
            # Clean faculty and year data
            data['Clean_Faculty'] = data['Faculty'].apply(self.clean_faculty_name)
            data['Clean_Year'] = data['Course Year'].apply(self.clean_year_name)
            
            # Parse industry preferences
            data['Industry_Numbers'] = data['Industries'].apply(self.parse_industry_numbers)
            
            # Create expanded dataset (one row per industry preference)
            expanded_data = []
            for _, row in data.iterrows():
                if row['Industry_Numbers']:  # If student has industry preferences
                    for industry_num in row['Industry_Numbers']:
                        expanded_data.append({
                            'Faculty': row['Clean_Faculty'],
                            'Year': row['Clean_Year'],
                            'Industry_Number': industry_num,
                            'Industry_Name': self.industry_mapping.get(industry_num, f"Unknown_{industry_num}"),
                            'Student_Count': 1
                        })
            
            expanded_df = pd.DataFrame(expanded_data)
            
            # Create pivot table
            pivot_table = pd.pivot_table(
                expanded_df,
                values='Student_Count',
                index='Industry_Name',
                columns=['Faculty', 'Year'],
                aggfunc='sum',
                fill_value=0,
                margins=True,
                margins_name='Total'
            )
            
            # Flatten column names for easier handling
            pivot_table.columns = [f"{faculty}_{year}" if year != 'Total' else 'Total' 
                                 for faculty, year in pivot_table.columns]
            
            self.analysis_results['pivot_table'] = pivot_table
            self.analysis_results['expanded_data'] = expanded_df
            
            print("Industry preferences table created successfully")
            return pivot_table
            
        except Exception as e:
            print(f"Error creating industry preferences table: {e}")
            return None
    
    def create_focused_table(self):
        """Create the focused table for Engineering and Arts faculties as requested"""
        if 'expanded_data' not in self.analysis_results:
            print("No expanded data available.")
            return None
            
        try:
            expanded_df = self.analysis_results['expanded_data']
            
            # Filter for Engineering and Arts faculties
            focused_faculties = ['Faculty of Engineering', 'Faculty of Arts and Social Sciences']
            focused_data = expanded_df[expanded_df['Faculty'].isin(focused_faculties)]
            
            # Create pivot table
            pivot_table = pd.pivot_table(
                focused_data,
                values='Student_Count',
                index='Industry_Name',
                columns=['Faculty', 'Year'],
                aggfunc='sum',
                fill_value=0
            )
            
            # Reorder columns to match requested format
            year_order = ['1st Year', '2nd Year', '3rd Year', '4th Year', '5th Year']
            faculty_order = ['Faculty of Engineering', 'Faculty of Arts and Social Sciences']
            
            # Create new column structure
            new_columns = []
            for faculty in faculty_order:
                for year in year_order:
                    if (faculty, year) in pivot_table.columns:
                        new_columns.append((faculty, year))
            
            if new_columns:
                pivot_table = pivot_table[new_columns]
            
            # Reorder rows to match the requested industry list
            industry_order = [
                'Accounting', 'Advertising, Media, Journalism, and Communications',
                'Agriculture and Environment', 'Animals and Vet', 'Architecture',
                'Arts, Humanities, and Politics', 'Building and Construction',
                'Business and Commerce', 'Community and Social Work',
                'Creative Arts and Music', 'Design', 'Economics and Finance',
                'Education, Childcare and Teaching', 'Engineering', 'Entrepreneur',
                'Food and Beverage', 'Government, Defence and Policing',
                'Hair and Beauty', 'Health and Sport Sciences', 'Law',
                'Marketing and Public Relations', 'Mathematics',
                'Medical Sciences and Medicine', 'Nursing and Midwifery',
                'Property and Real Estate', 'Psychology', 'Science', 'Technology'
            ]
            
            # Filter to only include industries that exist in the data
            existing_industries = [ind for ind in industry_order if ind in pivot_table.index]
            pivot_table = pivot_table.reindex(existing_industries)
            
            self.analysis_results['focused_table'] = pivot_table
            
            print("Focused industry preferences table created successfully")
            return pivot_table
            
        except Exception as e:
            print(f"Error creating focused table: {e}")
            return None
    
    def create_excel_report(self, output_filename="Industry_Preferences_Analysis.xlsx"):
        """Create Excel report with analysis results"""
        try:
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create Industry Mapping sheet
            self.create_industry_mapping_sheet(wb)
            
            # Create Focused Table sheet (Engineering vs Arts)
            self.create_focused_table_sheet(wb)
            
            # Create Full Analysis sheet
            self.create_full_analysis_sheet(wb)
            
            # Create Raw Data sheet
            self.create_raw_data_sheet(wb)
            
            # Save workbook
            wb.save(output_filename)
            print(f"Excel report saved as: {output_filename}")
            return True
            
        except Exception as e:
            print(f"Error creating Excel report: {e}")
            return False
    
    def create_industry_mapping_sheet(self, wb):
        """Create industry mapping reference sheet"""
        ws = wb.create_sheet("Industry Mapping")
        
        # Title
        ws.merge_cells('A1:B1')
        ws['A1'] = "INDUSTRY NUMBER TO NAME MAPPING"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        ws['A3'] = "Industry Number"
        ws['B3'] = "Industry Name"
        
        # Style headers
        for col in ['A3', 'B3']:
            ws[col].font = Font(bold=True)
            ws[col].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add mapping data
        row = 4
        for num, name in sorted(self.industry_mapping.items()):
            ws.cell(row=row, column=1, value=num)
            ws.cell(row=row, column=2, value=name)
            row += 1
        
        # Auto-adjust columns
        self.auto_adjust_columns(ws)
    
    def create_focused_table_sheet(self, wb):
        """Create the focused table sheet (Engineering vs Arts)"""
        ws = wb.create_sheet("Industry Preferences - Engineering vs Arts")
        
        # Title
        ws.merge_cells('A1:L1')
        ws['A1'] = "INDUSTRY PREFERENCES BY FACULTY AND YEAR GROUP"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        if 'focused_table' in self.analysis_results:
            pivot_table = self.analysis_results['focused_table']
            
            # Create headers
            headers = ['Industry']
            year_order = ['1st Year', '2nd Year', '3rd Year', '4th Year', '5th Year']
            faculty_order = ['Faculty of Engineering', 'Faculty of Arts and Social Sciences']
            
            for faculty in faculty_order:
                for year in year_order:
                    if (faculty, year) in pivot_table.columns:
                        headers.append(f"{faculty} - {year}")
            
            # Add headers
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=3, column=col_idx, value=header)
                ws.cell(row=3, column=col_idx).font = Font(bold=True)
                ws.cell(row=3, column=col_idx).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                ws.cell(row=3, column=col_idx).alignment = Alignment(horizontal='center')
            
            # Add data
            for row_idx, (industry, row) in enumerate(pivot_table.iterrows(), 4):
                ws.cell(row=row_idx, column=1, value=industry)
                ws.cell(row=row_idx, column=1).font = Font(bold=True)
                
                col_idx = 2
                for faculty in faculty_order:
                    for year in year_order:
                        if (faculty, year) in pivot_table.columns:
                            value = row[(faculty, year)]
                            ws.cell(row=row_idx, column=col_idx, value=value)
                            col_idx += 1
        
        # Auto-adjust columns
        self.auto_adjust_columns(ws)
    
    def create_full_analysis_sheet(self, wb):
        """Create full analysis sheet"""
        ws = wb.create_sheet("Full Industry Analysis")
        
        # Title
        ws.merge_cells('A1:Z1')
        ws['A1'] = "COMPLETE INDUSTRY PREFERENCES ANALYSIS"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        if 'pivot_table' in self.analysis_results:
            pivot_table = self.analysis_results['pivot_table']
            
            # Add pivot table data
            for row_idx, (index, row) in enumerate(pivot_table.iterrows(), 3):
                # Add row label
                ws.cell(row=row_idx, column=1, value=index)
                
                # Add data values
                for col_idx, value in enumerate(row, 2):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Add column headers
            for col_idx, col_name in enumerate(pivot_table.columns, 2):
                ws.cell(row=2, column=col_idx, value=col_name)
            
            # Add row header
            ws.cell(row=2, column=1, value="Industry")
            
            # Style headers
            for col in range(1, len(pivot_table.columns) + 2):
                cell = ws.cell(row=2, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust columns
        self.auto_adjust_columns(ws)
    
    def create_raw_data_sheet(self, wb):
        """Create raw data sheet"""
        ws = wb.create_sheet("Raw Data")
        
        if 'expanded_data' in self.analysis_results:
            expanded_df = self.analysis_results['expanded_data']
            
            # Add raw data
            for r in dataframe_to_rows(expanded_df, index=False, header=True):
                ws.append(r)
            
            # Style header
            self.style_header(ws, 1)
        
        # Auto-adjust columns
        self.auto_adjust_columns(ws)
    
    def style_header(self, ws, row):
        """Style the header row"""
        for cell in ws[row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
    
    def auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def print_analysis_results(self):
        """Print analysis results"""
        print("\n" + "="*80)
        print("INDUSTRY PREFERENCES ANALYSIS RESULTS")
        print("="*80)
        
        if 'focused_table' in self.analysis_results:
            print("\nFOCUSED TABLE - Engineering vs Arts Faculties:")
            print("-" * 80)
            focused_table = self.analysis_results['focused_table']
            print(focused_table.to_string())
        
        if 'expanded_data' in self.analysis_results:
            expanded_df = self.analysis_results['expanded_data']
            print(f"\nTotal industry preferences recorded: {len(expanded_df)}")
            print(f"Unique students with preferences: {expanded_df.groupby(['Faculty', 'Year']).size().sum()}")

def main():
    print("INDUSTRY PREFERENCES ANALYSIS")
    print("="*50)
    
    # Initialize analysis
    analysis = IndustryPreferencesAnalysis("August Export_SD 2 Sept.xlsx")
    
    # Load industry mapping
    if not analysis.load_industry_mapping():
        return
    
    # Load August data
    if not analysis.load_august_data():
        return
    
    # Create industry preferences table
    pivot_table = analysis.create_industry_preferences_table()
    if pivot_table is None:
        return
    
    # Create focused table
    focused_table = analysis.create_focused_table()
    if focused_table is None:
        return
    
    # Print results
    analysis.print_analysis_results()
    
    # Create Excel report
    if analysis.create_excel_report():
        print("\n" + "="*60)
        print("ANALYSIS COMPLETE")
        print("="*60)
        print("Excel report saved as: Industry_Preferences_Analysis.xlsx")
        print("This file contains:")
        print("1. Industry Mapping sheet - Number to name reference")
        print("2. Industry Preferences - Engineering vs Arts sheet - Main analysis table")
        print("3. Full Industry Analysis sheet - Complete analysis")
        print("4. Raw Data sheet - Expanded dataset")

if __name__ == "__main__":
    main()
