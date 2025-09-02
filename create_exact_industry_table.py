#!/usr/bin/env python3
"""
Create Exact Industry Preferences Table
This script creates the exact table format requested by the user,
showing industry preferences by faculty and year group.
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np
import re

class ExactIndustryTableCreator:
    def __init__(self, file_path):
        self.file_path = file_path
        self.industry_mapping = {}
        self.august_data = None
        
    def load_industry_mapping(self):
        """Load industry number to name mapping from Sheet 7"""
        try:
            # Read Sheet 7
            df = pd.read_excel(self.file_path, sheet_name='Sheet7', header=None)
            
            # Extract industry mappings (rows 2 onwards)
            for i in range(2, len(df)):
                row = df.iloc[i]
                if pd.notna(row[1]) and pd.notna(row[2]):  # Check if both number and name exist
                    industry_number = int(row[1])
                    industry_name = str(row[2]).strip()
                    self.industry_mapping[industry_number] = industry_name
            
            print(f"Loaded {len(self.industry_mapping)} industry mappings")
            return True
            
        except Exception as e:
            print(f"Error loading industry mapping: {e}")
            return False
    
    def load_august_data(self):
        """Load August data"""
        try:
            self.august_data = pd.read_excel(self.file_path, sheet_name='August')
            print(f"August data loaded: {len(self.august_data)} rows")
            return True
        except Exception as e:
            print(f"Error loading August data: {e}")
            return False
    
    def parse_industry_numbers(self, industry_string):
        """Parse industry numbers from string like '|14|15|12|27|28|'"""
        if pd.isna(industry_string):
            return []
        
        # Remove pipes and split by |
        numbers_str = str(industry_string).replace("'", "").replace("|", " ")
        numbers = re.findall(r'\d+', numbers_str)
        return [int(num) for num in numbers if int(num) in self.industry_mapping]
    
    def clean_faculty_name(self, faculty_string):
        """Clean faculty name by removing pipes and standardizing"""
        if pd.isna(faculty_string):
            return "Other"
        
        faculty = str(faculty_string).replace("'", "").replace("|", "").strip()
        
        # Standardize faculty names
        if "Faculty of Engineering" in faculty:
            return "Faculty of Engineering"
        elif "Faculty of Arts and Social Sciences" in faculty:
            return "Faculty of Arts and Social Sciences"
        else:
            return "Other"
    
    def clean_year_name(self, year_string):
        """Clean year name by removing pipes and standardizing"""
        if pd.isna(year_string):
            return "Unknown"
        
        year = str(year_string).replace("'", "").replace("|", "").strip()
        
        # Standardize year names
        if "1st Year" in year or year == "1":
            return "1st Year"
        elif "2nd Year" in year or year == "2":
            return "2nd Year"
        elif "3rd Year" in year or year == "3":
            return "3rd Year"
        elif "4th Year" in year or year == "4":
            return "4th Year"
        elif "5th Year" in year or year == "5":
            return "5th Year"
        else:
            return "Unknown"
    
    def create_exact_table(self):
        """Create the exact table format requested by the user"""
        if self.august_data is None:
            print("No August data loaded.")
            return None
            
        try:
            # Prepare data
            data = self.august_data.copy()
            
            # Clean faculty and year data
            data['Clean_Faculty'] = data['Faculty'].apply(self.clean_faculty_name)
            data['Clean_Year'] = data['Course Year'].apply(self.clean_year_name)
            
            # Parse industry preferences
            data['Industry_Numbers'] = data['Industries'].apply(self.parse_industry_numbers)
            
            # Create expanded dataset (one row per industry preference)
            expanded_data = []
            for _, row in data.iterrows():
                if row['Industry_Numbers']:  # If student has industry preferences
                    for industry_num in row['Industry_Numbers']:
                        expanded_data.append({
                            'Faculty': row['Clean_Faculty'],
                            'Year': row['Clean_Year'],
                            'Industry_Number': industry_num,
                            'Industry_Name': self.industry_mapping.get(industry_num, f"Unknown_{industry_num}"),
                            'Student_Count': 1
                        })
            
            expanded_df = pd.DataFrame(expanded_data)
            
            # Filter for only Engineering and Arts faculties
            focused_faculties = ['Faculty of Engineering', 'Faculty of Arts and Social Sciences']
            focused_data = expanded_df[expanded_df['Faculty'].isin(focused_faculties)]
            
            # Create pivot table
            pivot_table = pd.pivot_table(
                focused_data,
                values='Student_Count',
                index='Industry_Name',
                columns=['Faculty', 'Year'],
                aggfunc='sum',
                fill_value=0
            )
            
            # Reorder columns to match requested format
            year_order = ['1st Year', '2nd Year', '3rd Year', '4th Year', '5th Year']
            faculty_order = ['Faculty of Engineering', 'Faculty of Arts and Social Sciences']
            
            # Create new column structure
            new_columns = []
            for faculty in faculty_order:
                for year in year_order:
                    if (faculty, year) in pivot_table.columns:
                        new_columns.append((faculty, year))
            
            if new_columns:
                pivot_table = pivot_table[new_columns]
            
            # Reorder rows to match the requested industry list exactly
            industry_order = [
                'Accounting', 'Advertising, Media, Journalism, and Communications',
                'Agriculture and Environment', 'Animals and Vet', 'Architecture',
                'Arts, Humanities, and Politics', 'Building and Construction',
                'Business and Commerce', 'Community and Social Work',
                'Creative Arts and Music', 'Design', 'Economics and Finance',
                'Education, Childcare and Teaching', 'Engineering', 'Entrepreneur',
                'Food and Beverage', 'Government, Defence and Policing',
                'Hair and Beauty', 'Health and Sport Sciences', 'Law',
                'Marketing and Public Relations', 'Mathematics',
                'Medical Sciences and Medicine', 'Nursing and Midwifery',
                'Property and Real Estate', 'Psychology', 'Science', 'Technology',
                'Trades and Mining', 'Sports', 'Transport, Tourism and Hospitality',
                'Fashion', 'Australian Defence Force', 'Energy'
            ]
            
            # Create final table with all industries (even if no data)
            final_table = pd.DataFrame(index=industry_order)
            
            # Add data for existing columns
            for col in pivot_table.columns:
                final_table[col] = pivot_table[col]
            
            # Fill NaN with 0
            final_table = final_table.fillna(0)
            
            return final_table
            
        except Exception as e:
            print(f"Error creating exact table: {e}")
            return None
    
    def create_excel_table(self, table_data, output_filename="Exact_Industry_Preferences_Table.xlsx"):
        """Create Excel file with the exact table format"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Industry Preferences Table"
            
            # Title
            ws.merge_cells('A1:L1')
            ws['A1'] = "INDUSTRY PREFERENCES BY FACULTY AND YEAR GROUP"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Faculty headers
            ws.merge_cells('B2:F2')
            ws['B2'] = "Faculty of Engineering"
            ws.merge_cells('G2:K2')
            ws['G2'] = "Faculty of Arts and Social Sciences"
            
            # Style faculty headers
            for cell in ['B2', 'G2']:
                ws[cell].font = Font(bold=True, size=12)
                ws[cell].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                ws[cell].alignment = Alignment(horizontal='center')
            
            # Year headers
            year_headers = ['1st Year', '2nd Year', '3rd Year', '4th Year', '5th Year']
            for i, year in enumerate(year_headers):
                # Engineering years
                ws.cell(row=3, column=i+2, value=year)
                # Arts years
                ws.cell(row=3, column=i+7, value=year)
            
            # Style year headers
            for col in range(2, 12):
                cell = ws.cell(row=3, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Industry column header
            ws['A3'] = "Industry"
            ws['A3'].font = Font(bold=True)
            ws['A3'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            ws['A3'].alignment = Alignment(horizontal='center')
            
            # Add data
            for row_idx, (industry, row) in enumerate(table_data.iterrows(), 4):
                # Industry name
                ws.cell(row=row_idx, column=1, value=industry)
                ws.cell(row=row_idx, column=1).font = Font(bold=True)
                
                # Engineering data
                for col_idx, year in enumerate(year_headers):
                    col_name = ('Faculty of Engineering', year)
                    if col_name in table_data.columns:
                        value = table_data.loc[industry, col_name]
                    else:
                        value = 0
                    ws.cell(row=row_idx, column=col_idx+2, value=value)
                
                # Arts data
                for col_idx, year in enumerate(year_headers):
                    col_name = ('Faculty of Arts and Social Sciences', year)
                    if col_name in table_data.columns:
                        value = table_data.loc[industry, col_name]
                    else:
                        value = 0
                    ws.cell(row=row_idx, column=col_idx+7, value=value)
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
            
            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(output_filename)
            print(f"Excel table saved as: {output_filename}")
            return True
            
        except Exception as e:
            print(f"Error creating Excel table: {e}")
            return False
    
    def print_table(self, table_data):
        """Print the table in a formatted way"""
        print("\n" + "="*120)
        print("INDUSTRY PREFERENCES BY FACULTY AND YEAR GROUP")
        print("="*120)
        
        if table_data is not None:
            # Print faculty headers
            print(f"{'Industry':<40} {'Faculty of Engineering':<60} {'Faculty of Arts and Social Sciences':<60}")
            print(f"{'':<40} {'1st Year':<12} {'2nd Year':<12} {'3rd Year':<12} {'4th Year':<12} {'5th Year':<12} "
                  f"{'1st Year':<12} {'2nd Year':<12} {'3rd Year':<12} {'4th Year':<12} {'5th Year':<12}")
            print("-" * 120)
            
            # Print data
            for industry in table_data.index:
                print(f"{industry:<40}", end="")
                
                # Engineering data
                for year in ['1st Year', '2nd Year', '3rd Year', '4th Year', '5th Year']:
                    col_name = ('Faculty of Engineering', year)
                    if col_name in table_data.columns:
                        value = table_data.loc[industry, col_name]
                    else:
                        value = 0
                    print(f"{value:<12}", end="")
                
                # Arts data
                for year in ['1st Year', '2nd Year', '3rd Year', '4th Year', '5th Year']:
                    col_name = ('Faculty of Arts and Social Sciences', year)
                    if col_name in table_data.columns:
                        value = table_data.loc[industry, col_name]
                    else:
                        value = 0
                    print(f"{value:<12}", end="")
                
                print()  # New line

def main():
    print("CREATING EXACT INDUSTRY PREFERENCES TABLE")
    print("="*50)
    
    # Initialize creator
    creator = ExactIndustryTableCreator("August Export_SD 2 Sept.xlsx")
    
    # Load industry mapping
    if not creator.load_industry_mapping():
        return
    
    # Load August data
    if not creator.load_august_data():
        return
    
    # Create exact table
    table_data = creator.create_exact_table()
    if table_data is None:
        return
    
    # Print table
    creator.print_table(table_data)
    
    # Create Excel file
    if creator.create_excel_table(table_data):
        print("\n" + "="*60)
        print("TABLE CREATION COMPLETE")
        print("="*60)
        print("Excel table saved as: Exact_Industry_Preferences_Table.xlsx")
        print("This file contains the exact table format you requested with:")
        print("- All 34 industries listed")
        print("- Faculty of Engineering vs Faculty of Arts and Social Sciences")
        print("- Year groups: 1st Year, 2nd Year, 3rd Year, 4th Year, 5th Year")
        print("- Student counts for each combination")

if __name__ == "__main__":
    main()
