import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_july_august_comparison_sheet():
    """Create a new sheet comparing July and August emails with activity data"""
    
    try:
        file_path = "August_Export_SD_2_Sept_updated.xlsx"
        print(f"Creating July-August comparison sheet in: {file_path}")
        
        # Load the workbook
        workbook = load_workbook(file_path)
        
        # Read both sheets
        july_df = pd.read_excel(file_path, sheet_name='July ')
        august_df = pd.read_excel(file_path, sheet_name='August')
        
        print(f"July data shape: {july_df.shape}")
        print(f"August data shape: {august_df.shape}")
        
        # Clean column names
        july_df.columns = july_df.columns.str.strip()
        august_df.columns = august_df.columns.str.strip()
        
        # Create or get the comparison sheet
        if 'July_August_Email_Comparison' in workbook.sheetnames:
            workbook.remove(workbook['July_August_Email_Comparison'])
        
        comparison_sheet = workbook.create_sheet('July_August_Email_Comparison')
        
        # Find existing users (emails that appear in both sheets)
        july_emails = set(july_df['Email'].str.lower())
        august_emails = set(august_df['Email'].str.lower())
        existing_emails = july_emails.intersection(august_emails)
        
        print(f"Total unique emails in July: {len(july_emails)}")
        print(f"Total unique emails in August: {len(august_emails)}")
        print(f"Existing users (emails in both): {len(existing_emails)}")
        
        # Create comparison dataframe for existing users
        comparison_data = []
        
        for email in existing_emails:
            # Get July data
            july_row = july_df[july_df['Email'].str.lower() == email].iloc[0]
            # Get August data
            august_row = august_df[august_df['Email'].str.lower() == email].iloc[0]
            
            # Extract values for comparison
            july_login_count = july_row['Login Count']
            august_web_sessions = august_row['Web sessions']
            
            july_avg_time = july_row['Avg Login Time']
            august_avg_time = august_row['Avg Login Time']
            
            july_vwe = july_row['Virtual Work Experience']
            august_vwe = august_row['Virtual Work Experience']
            
            # Calculate increases
            login_increase = august_web_sessions - july_login_count
            time_increase = august_avg_time - july_avg_time
            vwe_increase = august_vwe - july_vwe if pd.notna(august_vwe) and pd.notna(july_vwe) else np.nan
            
            comparison_data.append({
                'Email': email,
                'First_Name_July': july_row['First name'],
                'First_Name_August': august_row['First name'],
                'July_Login_Count': july_login_count,
                'August_Web_Sessions': august_web_sessions,
                'Login_Increase': login_increase,
                'July_Avg_Login_Time': july_avg_time,
                'August_Avg_Login_Time': august_avg_time,
                'Time_Increase_Seconds': time_increase,
                'July_VWE': july_vwe,
                'August_VWE': august_vwe,
                'VWE_Increase': vwe_increase,
                'July_Person_Tag': july_row['Person tag'],
                'August_Person_Tag': august_row['Person tag'],
                'July_Industries': july_row['Industries'],
                'August_Industries': august_row['Industries'],
                'Career_Profiling_Flag': august_row['Career_Profiling_Flag'] if 'Career_Profiling_Flag' in august_row else 0
            })
        
        # Create comparison dataframe
        comparison_df = pd.DataFrame(comparison_data)
        
        print(f"Comparison data shape: {comparison_df.shape}")
        
        # Add headers to the sheet
        headers = [
            'Email', 'First Name (July)', 'First Name (August)',
            'July Login Count', 'August Web Sessions', 'Login Increase',
            'July Avg Login Time', 'August Avg Login Time', 'Time Increase (seconds)',
            'July VWE', 'August VWE', 'VWE Increase',
            'July Person Tag', 'August Person Tag',
            'July Industries', 'August Industries',
            'Career Profiling Flag'
        ]
        
        # Style definitions
        title_font = Font(bold=True, size=16, color="FFFFFF")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        data_font = Font(size=10)
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        neutral_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Add title
        comparison_sheet.merge_cells('A1:Q1')
        title_cell = comparison_sheet['A1']
        title_cell.value = "JULY-AUGUST EMAIL COMPARISON - EXISTING USERS ACTIVITY ANALYSIS"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = comparison_sheet.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data rows
        for row_idx, (_, row_data) in enumerate(comparison_df.iterrows(), 4):
            for col_idx, (key, value) in enumerate(row_data.items(), 1):
                cell = comparison_sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                
                # Apply conditional formatting for increase columns
                if col_idx == 6:  # Login Increase
                    if pd.notna(value) and value > 0:
                        cell.fill = positive_fill
                    elif pd.notna(value) and value < 0:
                        cell.fill = negative_fill
                    elif pd.notna(value) and value == 0:
                        cell.fill = neutral_fill
                
                elif col_idx == 9:  # Time Increase
                    if pd.notna(value) and value > 0:
                        cell.fill = positive_fill
                    elif pd.notna(value) and value < 0:
                        cell.fill = negative_fill
                    elif pd.notna(value) and value == 0:
                        cell.fill = neutral_fill
                
                elif col_idx == 12:  # VWE Increase
                    if pd.notna(value) and value > 0:
                        cell.fill = positive_fill
                    elif pd.notna(value) and value < 0:
                        cell.fill = negative_fill
                    elif pd.notna(value) and value == 0:
                        cell.fill = neutral_fill
        
        # Add summary section below the data
        summary_start_row = len(comparison_df) + 5
        
        # Summary title
        comparison_sheet.merge_cells(f'A{summary_start_row}:Q{summary_start_row}')
        summary_title = comparison_sheet[f'A{summary_start_row}']
        summary_title.value = "SUMMARY STATISTICS"
        summary_title.font = Font(bold=True, size=14, color="FFFFFF")
        summary_title.fill = title_fill
        summary_title.alignment = Alignment(horizontal='center', vertical='center')
        
        # Summary data
        summary_data = [
            ['Metric', 'Value', 'Formula', 'Description'],
            ['Total Existing Users', len(comparison_df), f'=COUNTA(A:A)-3', 'Count of users present in both July and August'],
            ['Users with Increased Login Activity', (comparison_df['Login_Increase'] > 0).sum(), '=COUNTIF(F:F,">0")', 'Number of users with more web sessions in August'],
            ['Users with Decreased Login Activity', (comparison_df['Login_Increase'] < 0).sum(), '=COUNTIF(F:F,"<0")', 'Number of users with fewer web sessions in August'],
            ['Users with Same Login Activity', (comparison_df['Login_Increase'] == 0).sum(), '=COUNTIF(F:F,0)', 'Number of users with same web sessions'],
            ['Users with Increased Login Time', (comparison_df['Time_Increase_Seconds'] > 0).sum(), '=COUNTIF(I:I,">0")', 'Number of users with longer login times in August'],
            ['Users with Decreased Login Time', (comparison_df['Time_Increase_Seconds'] < 0).sum(), '=COUNTIF(I:I,"<0")', 'Number of users with shorter login times in August'],
            ['Users with Increased VWE', (comparison_df['VWE_Increase'] > 0).sum(), '=COUNTIF(L:L,">0")', 'Number of users with increased VWE in August'],
            ['Users with Decreased VWE', (comparison_df['VWE_Increase'] < 0).sum(), '=COUNTIF(L:L,"<0")', 'Number of users with decreased VWE in August'],
            ['Average Login Increase', comparison_df['Login_Increase'].mean(), '=AVERAGE(F:F)', 'Average increase in web sessions from July to August'],
            ['Average Time Increase (seconds)', comparison_df['Time_Increase_Seconds'].mean(), '=AVERAGE(I:I)', 'Average increase in login time from July to August'],
            ['Average VWE Increase', comparison_df['VWE_Increase'].dropna().mean(), '=AVERAGE(L:L)', 'Average increase in VWE from July to August']
        ]
        
        for row_idx, row_data in enumerate(summary_data, summary_start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = comparison_sheet.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == summary_start_row + 1:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.font = data_font
                    if col_idx == 1:  # Metric column
                        cell.fill = neutral_fill
                        cell.font = Font(bold=True)
                    elif col_idx == 2:  # Value column
                        cell.font = Font(bold=True, color="0000FF")
                    elif col_idx == 3:  # Formula column
                        cell.font = Font(size=9, color="008000")
        
        # Add formula examples section
        formula_start_row = summary_start_row + len(summary_data) + 2
        
        comparison_sheet.merge_cells(f'A{formula_start_row}:Q{formula_start_row}')
        formula_title = comparison_sheet[f'A{formula_start_row}']
        formula_title.value = "FORMULA EXAMPLES FOR MANUAL CALCULATION"
        formula_title.font = Font(bold=True, size=14, color="FFFFFF")
        formula_title.fill = title_fill
        formula_title.alignment = Alignment(horizontal='center', vertical='center')
        
        formula_examples = [
            ['Calculation', 'Formula', 'Example', 'Result'],
            ['Login Increase', '=August_Web_Sessions - July_Login_Count', '=C2 - B2', 'Positive = Increased activity'],
            ['Time Increase', '=August_Avg_Login_Time - July_Avg_Login_Time', '=F2 - E2', 'Positive = Longer login time'],
            ['VWE Increase', '=August_VWE - July_VWE', '=I2 - H2', 'Positive = Increased VWE'],
            ['Percentage Change', '=(August_Value - July_Value)/July_Value', '=(C2-B2)/B2', 'Decimal result (multiply by 100 for %)']
        ]
        
        for row_idx, row_data in enumerate(formula_examples, formula_start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = comparison_sheet.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == formula_start_row + 1:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.font = data_font
                    if col_idx == 1:  # Calculation column
                        cell.fill = neutral_fill
                        cell.font = Font(bold=True)
                    elif col_idx == 2:  # Formula column
                        cell.font = Font(size=9, color="008000")
        
        # Auto-adjust column widths
        for column in comparison_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            comparison_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        workbook.save(file_path)
        print(f"July-August comparison sheet created successfully in: {file_path}")
        
        # Display summary
        print(f"\n📊 COMPARISON SHEET SUMMARY:")
        print(f"- Sheet name: 'July_August_Email_Comparison'")
        print(f"- Total existing users compared: {len(comparison_df)}")
        print(f"- Columns: {len(headers)} (including all comparison metrics)")
        print(f"- Conditional formatting applied for increases/decreases")
        print(f"- Summary statistics with formulas")
        print(f"- Formula examples for manual calculations")
        
        # Show sample data
        print(f"\n📋 SAMPLE COMPARISON DATA:")
        print(comparison_df[['Email', 'Login_Increase', 'Time_Increase_Seconds', 'VWE_Increase']].head())
        
        return comparison_df
        
    except Exception as e:
        print(f"Error creating comparison sheet: {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    comparison_df = create_july_august_comparison_sheet()
    if comparison_df is not None:
        print(f"\nJuly-August comparison sheet creation completed successfully!")
        print(f"Found {len(comparison_df)} existing users to compare")
    else:
        print(f"\nFailed to create comparison sheet.")
