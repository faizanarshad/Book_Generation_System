import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def update_august_data():
    """Update August data sheet by adding '1' in column T when 'Career Profiling Engaged' appears in column E"""
    
    try:
        # Read the Excel file
        file_path = "August Export_SD 2 Sept_modified.xlsx"
        print(f"Reading data from: {file_path}")
        
        # Load the workbook
        workbook = load_workbook(file_path)
        
        # Check if August sheet exists
        if 'August' not in workbook.sheetnames:
            print("Error: 'August' sheet not found!")
            return
        
        # Read the August sheet as DataFrame
        august_df = pd.read_excel(file_path, sheet_name='August')
        print(f"August sheet shape: {august_df.shape}")
        
        # Check current columns
        print(f"Current columns: {list(august_df.columns)}")
        
        # Check if column T already exists, if not create it
        if len(august_df.columns) < 20:  # Column T is the 20th column (index 19)
            # Add new column T
            august_df['Career_Profiling_Flag'] = 0
            print("Added new column 'Career_Profiling_Flag'")
        else:
            # Use existing column T
            august_df['Career_Profiling_Flag'] = 0
            print("Using existing column T")
        
        # Find rows where column E (Person tag) contains "Career Profiling Engaged"
        # Column E is index 4 (Person tag)
        person_tag_col = august_df.columns[4]  # Get the actual column name
        print(f"Checking column: {person_tag_col}")
        
        # Count matches before update
        matches_before = august_df[person_tag_col].str.contains('Career Profiling Engaged', na=False).sum()
        print(f"Found {matches_before} rows with 'Career Profiling Engaged' in {person_tag_col}")
        
        # Update column T (Career_Profiling_Flag) to 1 where condition is met
        mask = august_df[person_tag_col].str.contains('Career Profiling Engaged', na=False)
        august_df.loc[mask, 'Career_Profiling_Flag'] = 1
        
        # Count matches after update
        matches_after = august_df['Career_Profiling_Flag'].sum()
        print(f"Updated {matches_after} rows with '1' in Career_Profiling_Flag column")
        
        # Show sample of updated data
        print("\nSample of updated data:")
        updated_rows = august_df[august_df['Career_Profiling_Flag'] == 1]
        print(updated_rows[['First name', person_tag_col, 'Career_Profiling_Flag']].head())
        
        # Save to new file
        output_file = "August_Export_SD_2_Sept_updated.xlsx"
        
        # Create new workbook with updated data
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write August sheet with updates
            august_df.to_excel(writer, sheet_name='August', index=False)
            
            # Copy other sheets
            for sheet_name in workbook.sheetnames:
                if sheet_name != 'August':
                    sheet_df = pd.read_excel(file_path, sheet_name=sheet_name)
                    sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"\nUpdated data saved to: {output_file}")
        
        # Verify the update
        verification_df = pd.read_excel(output_file, sheet_name='August')
        verification_count = verification_df['Career_Profiling_Flag'].sum()
        print(f"Verification: {verification_count} rows have '1' in Career_Profiling_Flag column")
        
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
    except Exception as e:
        print(f"Error processing file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    update_august_data()
