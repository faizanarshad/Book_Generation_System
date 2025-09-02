#!/usr/bin/env python3
"""
Excel File Processor
This script can read Excel files, add formulas, and merge row data.
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import os

class ExcelProcessor:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        
    def load_excel(self):
        """Load the Excel file"""
        try:
            self.workbook = load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            print(f"Successfully loaded Excel file: {self.file_path}")
            return True
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            return False
    
    def read_data(self):
        """Read and display the current data in the Excel file"""
        if not self.workbook:
            print("No workbook loaded. Please load an Excel file first.")
            return None
            
        try:
            # Read with pandas for easier data manipulation
            df = pd.read_excel(self.file_path)
            print("Current data in Excel file:")
            print(df.head(10))  # Show first 10 rows
            print(f"\nTotal rows: {len(df)}")
            print(f"Columns: {list(df.columns)}")
            return df
        except Exception as e:
            print(f"Error reading data: {e}")
            return None
    
    def add_formula(self, cell_address, formula):
        """Add a formula to a specific cell"""
        if not self.worksheet:
            print("No worksheet loaded.")
            return False
            
        try:
            self.worksheet[cell_address] = formula
            print(f"Added formula '{formula}' to cell {cell_address}")
            return True
        except Exception as e:
            print(f"Error adding formula: {e}")
            return False
    
    def merge_rows_data(self, start_row, end_row, target_column, operation='sum'):
        """
        Merge data from multiple rows into a single cell
        
        Args:
            start_row: Starting row number (1-based)
            end_row: Ending row number (1-based)
            target_column: Column letter (e.g., 'A', 'B', 'C')
            operation: 'sum', 'average', 'concatenate', 'count'
        """
        if not self.worksheet:
            print("No worksheet loaded.")
            return False
            
        try:
            # Create formula based on operation
            if operation == 'sum':
                formula = f"=SUM({target_column}{start_row}:{target_column}{end_row})"
            elif operation == 'average':
                formula = f"=AVERAGE({target_column}{start_row}:{target_column}{end_row})"
            elif operation == 'count':
                formula = f"=COUNT({target_column}{start_row}:{target_column}{end_row})"
            elif operation == 'concatenate':
                # For concatenation, we'll need to use a different approach
                formula = f"=CONCATENATE({target_column}{start_row},\":\",{target_column}{end_row})"
            else:
                print(f"Unknown operation: {operation}")
                return False
            
            # Add the formula to a new row or specified location
            result_cell = f"{target_column}{end_row + 1}"
            self.worksheet[result_cell] = formula
            
            print(f"Added {operation} formula to cell {result_cell}: {formula}")
            return True
            
        except Exception as e:
            print(f"Error merging rows: {e}")
            return False
    
    def add_summary_formulas(self):
        """Add common summary formulas to the worksheet"""
        if not self.worksheet:
            print("No worksheet loaded.")
            return False
            
        try:
            # Get the last row and column
            max_row = self.worksheet.max_row
            max_col = self.worksheet.max_column
            
            # Add totals row
            for col in range(1, max_col + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                total_cell = f"{col_letter}{max_row + 1}"
                
                # Check if column contains numbers
                has_numbers = False
                for row in range(2, max_row + 1):  # Skip header row
                    cell_value = self.worksheet[f"{col_letter}{row}"].value
                    if isinstance(cell_value, (int, float)):
                        has_numbers = True
                        break
                
                if has_numbers:
                    # Add sum formula
                    self.worksheet[total_cell] = f"=SUM({col_letter}2:{col_letter}{max_row})"
                    print(f"Added SUM formula to {total_cell}")
            
            return True
            
        except Exception as e:
            print(f"Error adding summary formulas: {e}")
            return False
    
    def save_file(self, output_path=None):
        """Save the modified Excel file"""
        if not self.workbook:
            print("No workbook to save.")
            return False
            
        try:
            if output_path is None:
                # Create backup and save with modified name
                base_name = os.path.splitext(self.file_path)[0]
                output_path = f"{base_name}_modified.xlsx"
            
            self.workbook.save(output_path)
            print(f"File saved as: {output_path}")
            return True
            
        except Exception as e:
            print(f"Error saving file: {e}")
            return False

def main():
    # File path
    excel_file = "August Export_SD 2 Sept.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"Excel file '{excel_file}' not found in current directory.")
        return
    
    # Create processor instance
    processor = ExcelProcessor(excel_file)
    
    # Load the Excel file
    if not processor.load_excel():
        return
    
    # Read and display current data
    print("\n" + "="*50)
    print("CURRENT EXCEL FILE CONTENTS")
    print("="*50)
    data = processor.read_data()
    
    if data is not None:
        print("\n" + "="*50)
        print("ADDING FORMULAS AND MERGING DATA")
        print("="*50)
        
        # Example: Add summary formulas
        processor.add_summary_formulas()
        
        # Example: Add specific formulas (you can modify these)
        # processor.add_formula("A10", "=SUM(A1:A9)")
        # processor.add_formula("B10", "=AVERAGE(B1:B9)")
        
        # Example: Merge rows data
        # processor.merge_rows_data(1, 5, "A", "sum")
        # processor.merge_rows_data(1, 5, "B", "average")
        
        # Save the modified file
        processor.save_file()
        
        print("\n" + "="*50)
        print("PROCESSING COMPLETE")
        print("="*50)
        print("The modified Excel file has been saved with formulas and merged data.")
        print("You can now open the file to see the changes.")

if __name__ == "__main__":
    main()
